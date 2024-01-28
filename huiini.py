#-*- encoding: utf-8 -*-
from PySide2.QtCore import *
from PySide2.QtCore import Qt, QDir
from PySide2.QtGui import *
from PySide2 import QtGui, QtCore, QtWidgets
from PySide2.QtWidgets import QTableView, QTableWidget, QLineEdit, QTableWidgetItem, QFileDialog, QProgressDialog, QMessageBox, QListView, QAbstractItemView, QTreeView, QDialog, QVBoxLayout, QDialogButtonBox, QFileSystemModel, QInputDialog
from PySide2.QtWidgets import QPushButton, QListWidget, QListWidgetItem, QComboBox, QMenu, QAction
import sys
import guiV4
import cryptoDialog
from os import listdir, environ
from os.path import isfile, join, basename
import shutil
import os
try:
    import win32print
    import win32api
except:
    print("soy linux")
import time as time_old
from subprocess import Popen, call
from FacturasLocal import FacturaLocal as Factura
import math
import json
#import xlsxwriter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from openpyxl import load_workbook,  Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers, colors
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule


import openpyxl
from openpyxl.styles.alignment import Alignment

from datetime import datetime
from copy import copy
try:
    import ghostscript
except:
    print("ghostscript no está instalado")
import locale

import filecmp
import xlrd
from cryptography.fernet import Fernet


# import subprocess
# import psutil
# import signal


##C:\Python36\Scripts\pyside2-uic.exe mainwindowV2.ui -o guiV2.py
##C:\Python36\Scripts\pyside2-uic.exe mainwindowV4.ui -o guiV4.py
##C:\Python36\Scripts\pyinstaller.exe huiini.py
## C:\Users\Mio\AppData\Local\Packages\PythonSoftwareFoundation.Python.3.9_qbz5n2kfra8p0\LocalCache\local-packages\Python39\Scripts\pyinstaller.exe huiini.py
## C:\Users\Mio\AppData\Local\Packages\PythonSoftwareFoundation.Python.3.9_qbz5n2kfra8p0\LocalCache\local-packages\Python39\Scripts\pyside2-uic.exe mainwindowV4.ui -o guiV4.py

if getattr(sys, 'frozen', False):
    # we are running in a bundle
    scriptDirectory = os.path.dirname(sys.executable)
    appDataDirectory = os.path.expandvars('%APPDATA%\huiini')
else:
    # we are running in a normal Python environment
    scriptDirectory = os.path.dirname(os.path.abspath(__file__))
    appDataDirectory = join(scriptDirectory,"huiini_aux_files")






class PandasModel(QAbstractTableModel):
    """A model to interface a Qt view with pandas dataframe """

    def __init__(self, dataframe: pd.DataFrame, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._dataframe = dataframe

    def rowCount(self, parent=QModelIndex()) -> int:
        """ Override method from QAbstractTableModel

        Return row count of the pandas DataFrame
        """
        if parent == QModelIndex():
            return len(self._dataframe)

        return 0

    def columnCount(self, parent=QModelIndex()) -> int:
        """Override method from QAbstractTableModel

        Return column count of the pandas DataFrame
        """
        if parent == QModelIndex():
            return len(self._dataframe.columns)
        return 0

    def data(self, index: QModelIndex, role=Qt.ItemDataRole):
        """Override method from QAbstractTableModel

        Return data cell from the pandas DataFrame
        """
        if not index.isValid():
            return None

        if role == Qt.DisplayRole:
            return str(self._dataframe.iloc[index.row(), index.column()])

        return None

    def headerData(
        self, section: int, orientation: Qt.Orientation, role: Qt.ItemDataRole
    ):
        """Override method from QAbstractTableModel

        Return dataframe index as vertical header data and columns as horizontal header data.
        """
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return str(self._dataframe.columns[section])

            if orientation == Qt.Vertical:
                return str(self._dataframe.index[section])

        return None


class categorias_widget(QDialog):
    def __init__(self, parent=None):
        super(categorias_widget, self).__init__(parent)
        import ctypes
        user32 = ctypes.windll.user32
        user32.SetProcessDPIAware()
        screensize = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
        print(screensize)
        h = min(round(screensize[1]*0.8),850)
        self.setMinimumSize(520, h)
        layout = QVBoxLayout()
        self.add_button = QPushButton("Nueva")
        layout.addWidget(self.add_button)
        self.edit_button = QPushButton("Editar")
        layout.addWidget(self.edit_button)
        self.remove_button = QPushButton("Eliminar")
        layout.addWidget(self.remove_button)
        self.myListWidget = QListWidget()
        layout.addWidget(self.myListWidget)
        self.setLayout(layout)

class impresoras_widget(QDialog):
    def __init__(self, parent=None):
        super(impresoras_widget, self).__init__(parent)
        self.setMinimumSize(200, 100)
        layout = QVBoxLayout()
        self.lista = QListWidget()
        layout.addWidget(self.lista)
        self.setLayout(layout)
        self.lista.currentItemChanged.connect(self.cambiaSeleccionDeImpresora)

    def cambiaSeleccionDeImpresora(self, curr, prev):
        print(curr.text())
        self.impresoraDefault = curr.text()
        win32print.SetDefaultPrinter(self.impresoraDefault)

class getFilesDlg(QDialog):

    # sendPaths is a signal emitted by getPaths containing a list of file paths from
    # the users selection via this dialog.
    sendPaths = Signal(list)

    def __init__(self, parent=None):
        super(getFilesDlg, self).__init__(parent)
        import ctypes
        user32 = ctypes.windll.user32
        user32.SetProcessDPIAware()
        screensize = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
        print(screensize)
        h = min(round(screensize[1]*0.8),850)
        self.setMinimumSize(520,h)

        self.fileDlgPaths = []

        layout = QVBoxLayout()

        self.btnBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)

        self.btnBox.accepted.connect(self.getPaths)
        self.btnBox.rejected.connect(self.close)

        self.fsModel = QFileSystemModel()
        self.fsModel.setFilter(QtCore.QDir.AllDirs | QtCore.QDir.NoDotAndDotDot)

        self.treeView = QTreeView()
        self.treeView.setSelectionMode(QTreeView.ExtendedSelection)

        self.treeView.setModel(self.fsModel)
        self.treeView.setColumnWidth(0, 361)
        self.treeView.setColumnHidden(1, True)
        self.treeView.setColumnHidden(2, True)
        self.fsModel.setRootPath("")
        self.treeView.setSortingEnabled(True)
        layout.addWidget(self.treeView)
        layout.addWidget(self.btnBox)
        self.setLayout(layout)
        self.treeView.setRootIndex(self.fsModel.index(""))
        
        
        self.huiini_home_folder_path = ""
        if os.path.exists(os.path.join(appDataDirectory,"huiini_home_folder_path.txt")):
            with open(os.path.join(appDataDirectory,"huiini_home_folder_path.txt")) as f:
                self.huiini_home_folder_path = f.readline()
        print(self.huiini_home_folder_path)


        path_restante = self.huiini_home_folder_path
        while path_restante != os.path.split(path_restante)[0]:
            index_previo = self.fsModel.index(path_restante)
            self.treeView.expand(index_previo)
            path_restante = os.path.split(path_restante)[0]
        index_previo = self.fsModel.index(path_restante)
        self.treeView.expand(index_previo)

        # self.treeView.setRootIndex(self.fsModel.index("C:\\Dropbox"))
        # self.treeView.expand(self.treeView.rootIndex())
        #self.fsModel.setRootPath(environ['HOMEPATH'])
        # self.treeView.setRootIndex(self.fsModel.index("\\"))
        #self.treeView.expand(self.treeView.rootIndex())


    def getPaths(self):
    	# For some reason duplicates were being returned when they weren't supposed to.
    	# This obtains the selected files from the dialog and only returns individual
    	# paths.
        indexes = self.treeView.selectedIndexes()
        if indexes:
            self.fileDlgPaths = []
            for i in indexes:

    			# Possible permission error occuring here
    			# unable to replicate at this time
                path = self.fsModel.filePath(i)
                if path not in self.fileDlgPaths:
                    self.fileDlgPaths.append(path)
            self.close() # To close the dialog on an accept signal
            self.sendPaths.emit(self.fileDlgPaths)

class ImgWidgetPalomita(QtWidgets.QLabel):

    def __init__(self, parent=None):
        super(ImgWidgetPalomita, self).__init__(parent)
        pic_palomita = QtGui.QPixmap(join(scriptDirectory,"palomita.png"))
        self.setPixmap(pic_palomita)

class ImgWidgetTache(QtWidgets.QLabel):

    def __init__(self, parent=None):
        super(ImgWidgetTache, self).__init__(parent)
        pic_tache = QtGui.QPixmap(join(scriptDirectory,"x.png"))
        self.setPixmap(pic_tache)


class CryptoDialog(QtWidgets.QDialog, cryptoDialog.Ui_Dialog):
   
    def __init__(self, parent=None):
        super(CryptoDialog, self).__init__(parent)
        self.setupUi(self)
        self.saveButton.clicked.connect(self.guarda)
        self.cancelButton.clicked.connect(self.cierra)
        self.cliente_path = parent.cliente_path
        self.f = Fernet(parent.key)
        with open(join(parent.cliente_path,"Doc_Fiscal","claves"), 'rb') as fp:
            encrypted = fp.read()
            decrypted = self.f.decrypt(encrypted).decode()
            self.textEdit.setPlainText(decrypted) 
           

    def cierra(self):
        self.close()    
    def guarda(self):
        original = self.textEdit.toPlainText().encode()
        encrypted = self.f.encrypt(original)
        with open(join(self.cliente_path,"Doc_Fiscal","claves"), 'wb') as encrypted_file:
            encrypted_file.write(encrypted)
        self.cierra()


class Ui_MainWindow(QtWidgets.QMainWindow, guiV4.Ui_MainWindow):

    def __init__(self, parent=None):
        super(Ui_MainWindow, self).__init__(parent)
        self.setupUi(self)
        self.todos_los_meses = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]

        print(scriptDirectory)
        logoPix = QtGui.QPixmap(join(scriptDirectory,"logo_excel.png"))
        
        self.labelLogo.setPixmap(logoPix)

        logoSicadPix = QtGui.QPixmap(join(scriptDirectory,"logo_s.png"))
        self.labelLogo_sicad.setPixmap(logoSicadPix)

        
        with open(join(appDataDirectory,"conceptos.json"), "r") as jsonfile:
            self.concepto = json.load(jsonfile)

        with open(join(appDataDirectory,"kk.kk"), "rb") as keyfile:
            self.key = keyfile.read()

        with open(join(appDataDirectory,"cat_regimen.json"), "r") as jsonfile:
            self.regimen = json.load(jsonfile)

        self.tiene_pdflatex = True
        try:
            with open(os.path.join(appDataDirectory,"pdflatex_path.txt")) as f:
                self.pdflatex_path = f.readline()
        except:
            if shutil.which('pdflatex'):
                with open(os.path.join(appDataDirectory,"pdflatex_path.txt"), "w") as f:
                    f.write(shutil.which('pdflatex').replace("\\","\\\\"))
            else:
                reply = QMessageBox.question(self, 'No se detectó Miktex',"¿está Miktex instalado?\n contesta que si para buscar la ruta de pdflatex manualmete\n o no para cancelar", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

                if reply == QMessageBox.Yes:
                    path_to_file, _ = QFileDialog.getOpenFileName(self, "ruta de pdflatex", "~")
                    if "pdflatex.exe" in path_to_file.lower():
                        with open(os.path.join(appDataDirectory,"pdflatex_path.txt"), "w") as f:
                            f.write(path_to_file.replace("\\","\\\\"))
                    else:
                        self.warning(self, "Advertencia", "ruta incorrecta, la creación de pdfs quedará desactivada")
                        self.tiene_pdflatex = False
                if reply == QMessageBox.No:
                    self.warning(self, "Advertencia", "la creación de pdfs quedará desactivada")
                    self.tiene_pdflatex = False

        self.tiene_gswin64c = True
        try:
            with open(os.path.join(appDataDirectory,"gswin64c_path.txt")) as f:
                self.gswin64c_path = f.readline()
        except:
            if shutil.which('gswin64c'):
                with open(os.path.join(appDataDirectory,"gswin64c_path.txt"), "w") as f:
                    f.write(shutil.which('gswin64c').replace("\\","\\\\"))
                self.gswin64c_path = shutil.which('gswin64c').replace("\\","\\\\")
            else:
                reply = QMessageBox.question(self, 'No se detectó Ghostscript',"¿está Ghostscript instalado?\n contesta que si para buscar la ruta de gswin64c manualmete\n o no para cancelar", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

                if reply == QMessageBox.Yes:
                    path_to_file, _ = QFileDialog.getOpenFileName(self, "ruta de gswin64c", "~")
                    if "gswin64c.exe" in path_to_file.lower():
                        with open(os.path.join(appDataDirectory,"gswin64c_path.txt"), "w") as f:
                            f.write(path_to_file.replace("\\","\\\\"))
                        self.gswin64c_path = path_to_file.replace("\\","\\\\")
                    else:
                        self.warning(self, "Advertencia", "ruta incorrecta, la impresión quedará desactivada")
                        self.tiene_gswin64c = False
                if reply == QMessageBox.No:
                    self.warning(self, "Advertencia", "la impresión quedará desactivada")
                    self.tiene_gswin64c = False

        self.actionEscoger_cliente.triggered.connect(self.escoger_cliente)



        self.carpetaChooser.clicked.connect(self.cualCarpeta)
        self.action_editar_Categor_as.triggered.connect(self.edita_categorias)
        self.actionGenerar_Carpetas_Aspel_Coi.triggered.connect(self.carpetas_coi)
        self.actionClaves.triggered.connect(self.open_crypto)
        self.actionImprimir.triggered.connect(self.imprime)
        self.excel_anual_button.clicked.connect(self.abre_excel_anual)
        #self.descarga_bt.clicked.connect(self.descarga_mesta)

        self.actionSelccionar_Impresora.triggered.connect(self.cambiaImpresora)
        self.actionCancelar_Impresi_n.triggered.connect(self.cancelaImpresion)
        self.actionActualizar_cat_logos_CFDI.triggered.connect(self.actualizaCatalogos)

        if self.tiene_gswin64c == False:
            print("desabilitando la impresión....")
            self.actionCancelar_Impresi_n.setEnabled(False)
            self.actionSelccionar_Impresora.setEnabled(False)
            self.actionImprimir.setEnabled(False)

        self.tables = {}
        self.facturas = {}
        self.sumaRFC = {}





        #self.tableWidget_resumen.cellDoubleClicked.connect(self.meDoblePicaronResumen)
        self.progressBar.hide()

        self.tabWidget.currentChanged.connect(self.tabChanged)
        self.tabWidget.hide()
        self.numeroDeFacturasValidas = {}

    def open_crypto(self):
        widget = CryptoDialog(self)
        widget.exec()

    def actualizaCatalogos(self):
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        dialog.setNameFilter("Excel files (*.xlsx, *.xls)")
        dialog.setViewMode(QFileDialog.ViewMode.List)
        if dialog.exec():
            filenames = dialog.selectedFiles()
            
        filename = filenames[0]

        print(filename)

        
        book = xlrd.open_workbook(filename)
        
        sh = book.sheet_by_name("c_ClaveProdServ")
        print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))
        cat_conceptos = {}
        for row_idx in range(5, sh.nrows):
            key = str(sh.cell(row_idx, 0).value).split(".")[0]
            desc = str(sh.cell(row_idx, 1).value)
            cat_conceptos[key] = desc
            
        with open (join(appDataDirectory,"conceptos.json"), "w") as outfile:
            json.dump (cat_conceptos,outfile)

        self.concepto = cat_conceptos

        sh_unidad = book.sheet_by_name("c_ClaveUnidad")
        print("{0} {1} {2}".format(sh_unidad.name, sh_unidad.nrows, sh_unidad.ncols))
        cat_unidad = {}
        for row_idx in range(6, sh_unidad.nrows):
            key = str(sh_unidad.cell(row_idx, 0).value).split(".")[0]
            desc = str(sh_unidad.cell(row_idx, 1).value)
            cat_unidad[key] = desc
            
        with open (join(appDataDirectory,"cat_unidad.json"), "w") as outfile:
            json.dump (cat_unidad,outfile)

        sh_uso = book.sheet_by_name("c_UsoCFDI")
        print("{0} {1} {2}".format(sh_uso.name, sh_uso.nrows, sh_uso.ncols))
        cat_uso = {}
        for row_idx in range(6, sh_uso.nrows):
            key = str(sh_uso.cell(row_idx, 0).value).split(".")[0]
            desc = str(sh_uso.cell(row_idx, 1).value)
            cat_uso[key] = desc
            
        with open (join(appDataDirectory,"cat_uso.json"), "w") as outfile:
            json.dump (cat_uso,outfile)

        sh_regimen = book.sheet_by_name("c_RegimenFiscal")
        print("{0} {1} {2}".format(sh_regimen.name, sh_regimen.nrows, sh_regimen.ncols))
        cat_regimen = {}
        for row_idx in range(6, sh_regimen.nrows):
            key = str(sh_regimen.cell(row_idx, 0).value).split(".")[0]
            desc = str(sh_regimen.cell(row_idx, 1).value)
            cat_regimen[key] = desc
            
        with open (join(appDataDirectory,"cat_regimen.json"), "w") as outfile:
            json.dump (cat_regimen,outfile)





        QMessageBox.information(self, "Información", "Actualización de catálogos exitosa")
        

    def warning(self, parent, title, message):
        QMessageBox.information(parent, title, message)
        self.pluma.write(title)
        self.pluma.write("\n")
        self.pluma.write(message)
        self.pluma.write("\n")

    def carpetas_coi(self):
        print("carpetas coi")
        
        self.carpeta_this_year = join(self.cliente_path, str(datetime.now().year))
        self.folder_year = str(datetime.now().year)
        paths_meses = []
        carpetas_meses = ["01 ENERO", "02 FEBRERO", "03 MARZO", "04 ABRIL", "05 MAYO", "06 JUNIO", "07 JULIO", "08 AGOSTO", "09 SEPTIEMBRE", "10 OCTUBRE", "11 NOVIEMBRE", "12 DICIEMBRE"]
        for filename in os.listdir(self.carpeta_this_year):
            if os.path.isdir(join(self.carpeta_this_year,filename)):
                if filename in carpetas_meses:
                    print(filename)
                    paths_meses.append(join(self.carpeta_this_year,filename))

        for folder_mes in paths_meses:
            aspel_coi_path = join(folder_mes, "aspel_coi")
            if not os.path.isdir(aspel_coi_path):
                os.makedirs(aspel_coi_path)
            self.esteFolder = join(folder_mes,"EGRESOS")
            cuantosDuplicados = 0
            self.listaDeDuplicados = []
            self.listaDeFacturas = []
            self.listaDeUUIDs = []
            contador = 0
            for root, dirs, files in os.walk(self.esteFolder, topdown=False):
                for name in files:
                    if name.endswith(".xml") and not "CANCELA" in root:
                        try:
                            print(os.path.join(root, name))
                            laFactura = Factura(join(root, name))
                            if laFactura.sello == "SinSello":
                                print("Omitiendo xml sin sello "+laFactura.xml_path)
                            else:
                                if laFactura.version:
                                    if laFactura.UUID in self.listaDeUUIDs:
                                        cuantosDuplicados+=1
                                        self.listaDeDuplicados.append(laFactura.UUID)
                                    else:
                                        self.listaDeUUIDs.append({"uuid": laFactura.UUID,
                                                                "path": join(root, name),
                                                                "EmisorRFC": laFactura.EmisorRFC,
                                                                "ReceptorRFC": laFactura.ReceptorRFC})            
                        except:
                            print("falla sello")
            
            for f in self.listaDeUUIDs:
                oldfilename = os.path.basename(f["path"])
                try:
                    shutil.copy(f["path"],aspel_coi_path) 
                    new_dst_file_name = join(aspel_coi_path, f["EmisorRFC"]+"_"+oldfilename)
                    os.rename(join(aspel_coi_path, oldfilename), new_dst_file_name)#rename
                except:
                    print("este no es o ya estaba")
            


    def checkIfValuesExists1(self, dfObj, lista_d):
        resultDict = {}
        # Iterate over the list of elements one by one
        for fac in lista_d:
            
            # Check if the element exists in dataframe values
            if fac["uuid"] in dfObj.values:
                resultDict[fac["path"]] = True
            else:
                resultDict[fac["path"]] = False
        # Returns a dictionary of values & thier existence flag        
        return resultDict


    def escoger_cliente(self):
        file_dialog = getFilesDlg()
        file_dialog.sendPaths.connect(self.procesa_cliente)
        file_dialog.exec()

    def despliega_cliente(self):
        f = Fernet(self.key)
        if os.path.exists(join(self.cliente_path,"Doc_Fiscal","claves.txt")):
            #si existe el claves.txt (no encriptado) leerlo, crear el archivo encriptado y borrar el original
            with open(join(self.cliente_path,"Doc_Fiscal","claves.txt"), "rb") as fp:
                original = fp.read()
                encrypted = f.encrypt(original)
                with open(join(self.cliente_path,"Doc_Fiscal","claves"), 'wb') as encrypted_file:
                    encrypted_file.write(encrypted)

            os.remove(join(self.cliente_path,"Doc_Fiscal","claves.txt"))
        else:
            if not os.path.exists(join(self.cliente_path,"Doc_Fiscal","claves")):
                #si no hay ni claves.txt ni claves, preguntar nombre y rfc y crearlo encriptado
                os.makedirs(join(self.cliente_path,"Doc_Fiscal"), exist_ok=True)
                nombre, ok1 = QInputDialog.getText(self, 'Información del cliente faltante, ingresa el nombre', 'Nombre:')
                rfc, ok2 = QInputDialog.getText(self, 'Información del cliente faltante, ingresa el RFC', 'RFC:')
                with open(join(self.cliente_path,"Doc_Fiscal","claves"), 'wb') as encrypted_file:
                    info = "Nombre: "+nombre+"\n"+"RFC: "+rfc+"\n"
                    message = info.encode()
                    encrypted = f.encrypt(message)
                    encrypted_file.write(encrypted)
        
        #leer el encryptado, desencriptarlo y sacar nombre y rfc        
        with open(join(self.cliente_path,"Doc_Fiscal","claves"), 'rb') as fp:
            encrypted = fp.read()
            decrypted = f.decrypt(encrypted).decode()
            lines = decrypted.split("\n")
            self.nombre = ""
            self.rfc = ""
            for line in lines:
                if "Nombre: " in line:
                    self.nombre = line.split("Nombre: ")[1]
                if "RFC: " in line:
                    self.rfc = line.split("RFC: ")[1]       
        
        self.header_cliente.setText("Nombre: "+self.nombre+"\nRFC: "+self.rfc)
       
    def procesa_cliente(self,paths):
        self.cliente_path = paths.copy()[0]
        self.carpeta_this_year = join(self.cliente_path, str(datetime.now().year))
        self.folder_year = str(datetime.now().year)
        self.cargaCategorias()
        self.despliega_cliente()

        
        
        
        #############################################################
        # leer el excel anual de este año y hacer dataframes
        #############################################################
        import re
        
        
        cliente_str = os.path.basename(os.path.normpath(self.cliente_path))
        if os.path.exists(join(self.carpeta_this_year, cliente_str+"_"+str(datetime.now().year)+".xlsx")):
            self.dataFrames = {}
            self.annual_xlsx_path = join(self.carpeta_this_year, cliente_str+"_"+str(datetime.now().year)+".xlsx")
            workbook = load_workbook(self.annual_xlsx_path)
            for sheetname in workbook.sheetnames:
                print(sheetname)
                if sheetname in self.todos_los_meses:
                    print("facturas de "+sheetname+" en "+cliente_str+"_"+str(datetime.now().year)+".xlsx")
                    ws = workbook[sheetname]
                    data = ws.values
                    # Get the first line in file as a header line
                    columns = next(data)[0:]
                    # Create a DataFrame based on the second and subsequent lines of data
                    self.dataFrames[sheetname] = pd.DataFrame(data, columns=columns)

            los_meses = []
            for filename in os.listdir(self.carpeta_this_year):
                if os.path.isdir(join(self.carpeta_this_year,filename)):## aqui falta algo mas para asegurar que la carpeta sea un mes
                    print(filename)
                    los_meses.append(join(self.carpeta_this_year,filename))

            for folder_mes in los_meses:
                self.esteFolder = join(folder_mes,"EGRESOS")
                #el_mes = re.sub(r'\s*\d+\s*', '', os.path.basename(folder_mes)) 
                el_mes = self.extrae_mes(folder_mes) 
                print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>",el_mes)
                cuantosDuplicados = 0
                self.listaDeDuplicados = []
                self.listaDeFacturas = []
                self.listaDeUUIDs = []
                contador = 0
                for filename in os.listdir(self.esteFolder):
                    if filename.endswith(".xml"):
                        try:
                            print(join(self.esteFolder, filename))
                            laFactura = Factura(join(self.esteFolder + os.sep,filename))
                            if laFactura.sello == "SinSello":
                                print("Omitiendo xml sin sello "+laFactura.xml_path)
                            else:
                                if laFactura.version:
                                    if laFactura.UUID in self.listaDeUUIDs:

                                        cuantosDuplicados+=1
                                        self.listaDeDuplicados.append(laFactura.UUID)
                                    else:
                                        self.listaDeUUIDs.append({"uuid": laFactura.UUID, "path": join(self.esteFolder + os.sep,filename)})            
                        except:
                            print("falla")

                print(self.dataFrames[el_mes]) 
                print(self.listaDeUUIDs)
                result = self.checkIfValuesExists1(self.dataFrames[el_mes], self.listaDeUUIDs)
                print(result)
                message = "Se actualizará el mes de "+ el_mes + "\ncon las siguinetes facturas:"
                for key, value in result.items():
                    if not value:
                        message += "\n"+key
                if message != "Se actualizará el mes de "+ el_mes + "\ncon las siguinetes facturas:":
                    self.warning(self, "Actualización del excel anual", message)
                #self.annual_xlsx_path
                self.mes = el_mes
                self.agregaFacturasFaltantes(result)
                self.rellena_mes_gui_from_excel()
        
                
                #self.agregaTabMes(sheetname)
           
        else:
            print(self.carpeta_this_year)
            los_meses = []
            for filename in os.listdir(self.carpeta_this_year):
                if os.path.isdir(join(self.carpeta_this_year,filename)):
                    print(filename)
                    los_meses.append(join(self.carpeta_this_year,filename))

            #self.procesaCarpetas(los_meses)      
             
    def agregaFacturasFaltantes(self, result):
        
        for key, value in result.items():
            if not value:
                print("agregaria ", key)
                workbook = load_workbook(self.annual_xlsx_path)
                ws_todos = workbook["Conceptos"]
                ws_mes = workbook[self.mes]
                self.status_column = 0
                laFactura = Factura(key)

                row_mes = ws_mes.max_row+1

                dv = DataValidation(type="list", formula1='"Pendiente,Pagado"', allow_blank=True)
                ws_mes.add_data_validation(dv)

                

                ws_mes.cell(row_mes, 1, laFactura.conceptos[0]['clave_concepto'])
                ws_mes.cell(row_mes, 2, laFactura.fechaTimbrado)
                ws_mes.cell(row_mes, 3, laFactura.UUID)
                ws_mes.cell(row_mes, 4, laFactura.EmisorNombre)
                ws_mes.cell(row_mes, 5, laFactura.EmisorRFC)
                ws_mes.cell(row_mes, 6, laFactura.conceptos[0]['descripcion'])
                if laFactura.tipoDeComprobante == "E":
                    ws_mes.cell(row_mes, 7, 0.0 - laFactura.subTotal)
                    ws_mes.cell(row_mes, 8, 0.0 - laFactura.descuento)
                    ws_mes.cell(row_mes, 9, 0.0 - laFactura.traslados["IVA"]["importe"])
                    ws_mes.cell(row_mes, 10, 0.0)
                    ws_mes.cell(row_mes, 11, 0.0)
                    ws_mes.cell(row_mes, 12, 0.0)
                    ws_mes.cell(row_mes, 13, 0.0 - laFactura.total)
                else:
                    ws_mes.cell(row_mes, 7, laFactura.subTotal)
                    ws_mes.cell(row_mes, 8, laFactura.descuento)
                    ws_mes.cell(row_mes, 9, laFactura.traslados["IVA"]["importe"])
                    ws_mes.cell(row_mes, 10, laFactura.trasladosLocales["TUA"]["importe"])
                    ws_mes.cell(row_mes, 11, laFactura.trasladosLocales["ISH"]["importe"])
                    ws_mes.cell(row_mes, 12, laFactura.traslados["IEPS"]["importe"])
                    ws_mes.cell(row_mes, 13, laFactura.total)
                ws_mes.cell(row_mes, 14, laFactura.formaDePagoStr)
                ws_mes.cell(row_mes, 15, laFactura.metodoDePago)
                ws_mes.cell(row_mes, 16, laFactura.conceptos[0]['tipo'])
                status = "Pendiente"
                if laFactura.metodoDePago == "PUE":
                    status = "Pagado"
                # if laFactura.metodoDePago == "PPD": ############################# esto talvez hay que hacerlo mejor
                #     if laFactura.UUID in self.complementosDePago:
                #         if laFactura.total - self.complementosDePago[laFactura.UUID]["suma"] < 0.5:
                #             status = "Pagado"
                if laFactura.tipoDeComprobante == "P":
                    status = "Pagado"

                dv.add(ws_mes.cell(row_mes, 17))
                ws_mes.cell(row_mes, 17, status)
                ws_mes.cell(row_mes, 18, laFactura.tipoDeComprobante)
                # if laFactura.UUID in self.complementosDePago:
                #     ws_mes.cell(row_mes, 19, self.complementosDePago[laFactura.UUID]["suma"])

                if laFactura.tipoDeComprobante == "P":
                    print("segun "+ laFactura.UUID + "del mes " +self.mes+ ", aqui buscaria en todos los meses el uuid "+laFactura.IdDocumento+" y si encuentra su factura modificaria, la columna 13 del renglon de esa factura en el mes que esté, a Pagado")


                los_conceptos = laFactura.conceptos.copy()
                for concepto in los_conceptos:
                    concepto["subTotal"] = 0.0
                    concepto["mes"] = self.mes
                    concepto["UUID"] = laFactura.UUID
                    if concepto["impuestos"]:
                        concepto["impuestos"] = float(concepto['impuestos'])
                    else:
                        concepto["impuestos"] = 0
                    if laFactura.tipoDeComprobante == "E":
                        concepto["subTotal"] = float(concepto['importeConcepto']) - float(concepto['descuento'])
                        #concepto["importeConcepto"] = 0.0 - float(concepto['importeConcepto'])
                        concepto["descuento"] = 0.0 - float(concepto['descuento'])
                        #concepto["subTotal"] = 0.0 - float(concepto['subTotal'])
                        concepto["impuestos"] = 0.0 - float(concepto['impuestos'])
                        #concepto["total"] = 0.0 - float(concepto['total'])

                for i in range(1,ws_mes.max_column+1):
                    if ws_mes.cell(8, i).value == "Status":
                        self.status_column = i - 2

                row = ws_todos.max_row
                dv_categorias = DataValidation(type="list", formula1="=Categorias!A$1:A$"+str(len(self.lista_categorias_default)), allow_blank=True)
                ws_todos.add_data_validation(dv_categorias)
                for concepto in los_conceptos:
                    row += 1
                    clave = concepto['clave_concepto']
                    ws_todos.cell(row, 1, concepto['mes'])
                    ws_todos.cell(row, 2, clave)
                    ws_todos.cell(row, 3, self.getDescription(clave))
                    ws_todos.cell(row, 4, concepto['UUID'])
                    ws_todos.cell(row, 5, concepto['cantidad'])
                    ws_todos.cell(row, 6, concepto['descripcion'])
                    ws_todos.cell(row, 7, concepto['importeConcepto'])
                    ws_todos.cell(row, 8, concepto['descuento'])
                    ws_todos.cell(row, 9, concepto['importeConcepto'] - concepto['descuento'])
                    ws_todos.cell(row, 10, concepto['impuestos'])
                    ws_todos.cell(row, 11, (concepto['importeConcepto'] - concepto['descuento']) + concepto['impuestos'])
                    dv_categorias.add(ws_todos.cell(row, 12))
                    ws_todos.cell(row, 12, concepto['tipo'])
                    ws_todos.cell(row, 13, "=VLOOKUP(D"+str(row)+","+concepto['mes']+"!C:Q,"+str(self.status_column)+",FALSE)")
                workbook.save(self.annual_xlsx_path)

        

    def tabChanged(self, index):
        print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA estoy cambiando a ",str(index))
        name = self.tabWidget.tabText(index)
        folder_mes = {"ENERO":"01 ENERO",
                    "FEBRERO":"02 FEBRERO",
                    "MARZO":"03 MARZO",
                    "ABRIL":"04 ABRIL",
                    "MAYO":"05 MAYO",
                    "JUNIO":"06 JUNIO",
                    "JULIO":"07 JULIO",
                    "AGOSTO":"08 AGOSTO",
                    "SEPTIEMBRE":"09 SEPTIEMBRE",
                    "OCTUBRE":"10 OCTUBRE",
                    "NOVIEMBRE":"11 NOVIEMBRE",
                    "DICIEMBRE":"12 DICIEMBRE",
                    }
        
        
        
        if name == "Ingresos":
            print("aquí haría algo")
            self.mes = ""
        else:
            self.mes = name
            cliente = os.path.split(self.cliente_path)[1]
            self.excel_path = join(self.folder_year, folder_mes[name],"EGRESOS","huiini","resumen_" + cliente + "_" + self.mes + ".xlsx")
            n = -1
            lc = ["Excel"]

            for i in range(6,self.tables[self.mes].columnCount()):
                try: 
                    f = float(self.tables[self.mes].item(0,i).text())
                    header = self.tables[self.mes].horizontalHeaderItem(i).text()
                    lc.append(header)
                except:
                    print("noesnumero")
            
            
            #self.sumale()
        
    def setupTabMeses(self):
        self.tables[self.mes].setColumnCount(19)
        self.tables[self.mes].setColumnWidth(0,30)#pdf
        self.tables[self.mes].setColumnWidth(1,95)#fecha
        self.tables[self.mes].setColumnWidth(2,70)#uuid
        self.tables[self.mes].setColumnWidth(3,120)#receptor-nombre
        self.tables[self.mes].setColumnWidth(4,120)#emisor-rfc
        self.tables[self.mes].setColumnWidth(5,120)#concepto
        self.tables[self.mes].setColumnWidth(6,75)#Subtotal
        self.tables[self.mes].setColumnWidth(7,80)#Descuento
        self.tables[self.mes].setColumnWidth(8,80)#traslados-iva
        self.tables[self.mes].setColumnWidth(9,80)#traslados-ieps
        self.tables[self.mes].setColumnWidth(10,80)#traslados-ish
        self.tables[self.mes].setColumnWidth(11,80)#traslados-tua
        self.tables[self.mes].setColumnWidth(12,75)#retIVA
        self.tables[self.mes].setColumnWidth(13,75)#retISR
        self.tables[self.mes].setColumnWidth(14,80)#total
        self.tables[self.mes].setColumnWidth(15,74)#formaDePago
        self.tables[self.mes].setColumnWidth(16,77)#metodoDePago
        self.tables[self.mes].setColumnWidth(17,77)#tipo
        self.tables[self.mes].setColumnWidth(18,77)#carpetasCoi

        self.tables[self.mes].verticalHeader().setFixedWidth(35)
        header = self.tables[self.mes].verticalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self.handleHeaderMenu)

        lc = ["Pdf","Fecha","UUID","Receptor","Emisor","Concepto","Subtotal","Descuento","Traslado\nIVA","Traslado\nIEPS","Traslado\nISH", "Traslado\nTUA", "Retención\nIVA","Retención\nISR","Total","Forma\nPago","Método\nPago","Tipo","Carpetas Coi"]
        self.ponEncabezado(lc,self.mes)

        self.tables[self.mes].cellDoubleClicked.connect(self.meDoblePicaronXML)
        #self.tables[self.mes].horizontalHeader().sectionClicked.connect(self.reordena)
        self.tables[self.mes].setSortingEnabled(True)

#    def setupTabAnuales(self,tabName):


    def reordena(self, column):
        print("reodenaria")
        if column == 2:
            print("reodenaria con uuid")
            self.facturas[self.mes] = sorted(self.facturas[self.mes], key=lambda facturas: facturas.UUID)
        if column == 15:
            print("reodenaria con tipo")
            self.facturas[self.mes] = sorted(self.facturas[self.mes], key=lambda facturas: facturas.conceptos[0]['tipo'])

    def quitaCategoria(self):
        reply = QMessageBox.question(self, 'Message',"Estás seguro?", QMessageBox.Yes |
        QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            curr_indexes = self.cats_dialog.myListWidget.selectedIndexes()
            if len(curr_indexes)>1:
                print("no")
            else:
                #confirmacion para maricas
                categoria = self.cats_dialog.myListWidget.currentItem().text().split(" (")[0]
                self.dicc_de_categorias.pop(categoria)
                with open(self.json_path, "w", encoding="utf-8") as jsonfile:
                    json.dump(self.dicc_de_categorias, jsonfile, indent=4, sort_keys=True)

                self.cats_dialog.myListWidget.takeItem(curr_indexes[0].row())

    def editaCategoria(self):
        curr_indexes = self.cats_dialog.myListWidget.selectedIndexes()
        if len(curr_indexes)>1:
            print("no")
        else:
            categoria = self.cats_dialog.myListWidget.currentItem().text().split(" (")[0]
            lista_de_empiezos = ", ".join(self.dicc_de_categorias[categoria])


            nombre, ok1 = QInputDialog().getText(self.cats_dialog, "Nombre de la Categoría",
                                        "Nombre de la categoría:", QLineEdit.Normal,categoria)
            claves_ps, ok2 = QInputDialog().getText(self.cats_dialog, "Lista de claves de producto o servicio",
                                         "clave_ps:", QLineEdit.Normal,
                                         lista_de_empiezos)

            if ok1 and ok2:
                if nombre != categoria:
                    self.dicc_de_categorias.pop(categoria)
                if nombre in self.dicc_de_categorias:
                    #lista_previa = self.dicc_de_categorias[nombre].copy()

                    #lista_previa.extend(claves_ps.strip().split(","))#el espacio no se ocupa
                    self.dicc_de_categorias[nombre] = claves_ps.replace(" ","").split(",")
                #self.lista_ordenada = sorted(self.lista_ordenada, key=lambda tup: tup[1])
                with open(self.json_path, "w", encoding="utf-8") as jsonfile:
                    json.dump(self.dicc_de_categorias, jsonfile, indent=4, sort_keys=True)
                self.cats_dialog.myListWidget.clear()
                self.enlista_categorias()

    def agregaCategoria(self):
        nombre, ok1 = QInputDialog().getText(self.cats_dialog, "Nombre de la Categoría",
                                     "Nombre de la categoría:", QLineEdit.Normal,"")
        claves_ps, ok2 = QInputDialog().getText(self.cats_dialog, "Lista de claves de producto o servicio",
                                     "clave_ps:", QLineEdit.Normal,
                                     "")
        claves_ps.strip()

        if ok1 and ok2:
            for clave in claves_ps.split(", "):
                pasa = True
                for categoria, lista in self.dicc_de_categorias.items():
                    for clave1 in lista:
                        if clave1.startswith(clave) or clave.startswith(clave1):
                            pasa = False
                            QMessageBox.information(self, "Advertencia", "El inicio de clave " + clave + " ya está considerado en la categoría " + categoria)
                if pasa:
                    if clave == "" or nombre == "":
                        print("no mames")
                    else:
                        self.dicc_de_categorias[nombre] = claves_ps.split(", ")

            #self.lista_ordenada = sorted(self.lista_ordenada, key=lambda tup: tup[1])
            with open(self.json_path, "w", encoding="utf-8") as jsonfile:
                json.dump(self.dicc_de_categorias, jsonfile, indent=4, sort_keys=True)
            self.cats_dialog.myListWidget.clear()
            self.enlista_categorias()

    def enlista_categorias(self):
        for key, value in self.dicc_de_categorias.items():
            if len(value) > 3:
                texto_claves = " ("+value[0]+", "+value[1]+", "+value[2]+"...)"
            else:
                texto_claves = " (" + ", ".join(value) + ")"
            i = QListWidgetItem(key+texto_claves)
            i.setToolTip("\n".join(value))
            # if "Default" in tupla[0]:
            #     i.setBackground(QtGui.QColor("#ababab"))
            self.cats_dialog.myListWidget.addItem(i)

    def abre_excel_anual(self):
        try:
            os.startfile(self.annual_xlsx_path)
            print("este guey me pico:"+self.annual_xlsx_path)
        except:
            print ("el sistema no tiene una aplicacion por default para abrir exceles")
            QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir exceles" )
    def edita_categorias(self):
        self.cats_dialog = categorias_widget()
        self.cats_dialog.remove_button.clicked.connect(self.quitaCategoria)
        self.cats_dialog.add_button.clicked.connect(self.agregaCategoria)
        self.cats_dialog.edit_button.clicked.connect(self.editaCategoria)
        folder_cliente = os.path.split(os.path.split(self.paths[0])[0])[0]
        self.json_path = join(folder_cliente, "categorias_dicc_huiini.json")
        if os.path.exists(self.json_path):
            with open(self.json_path, "r", encoding="utf-8") as jsonfile:
                self.dicc_de_categorias = json.load(jsonfile)
        else:
            self.dicc_de_categorias = {}


        # lista_de_tuplas.extend(lista_categorias_default)
        #self.lista_ordenada = sorted(lista_de_tuplas, key=lambda tup: tup[1])

        self.enlista_categorias()
        self.cats_dialog.exec()

    def as_text(self,value):
        if value is None:
            return ""
        return str(value)

    def style_ws(self, ws, columna_totales, sumas_row):
        cell_border = Border(left=Side(border_style='medium', color='FF000000'),
                     right=Side(border_style='medium', color='FF000000'),
                     top=Side(border_style='medium', color='FF000000'),
                     bottom=Side(border_style='medium', color='FF000000'))

        cell_border_sumas = Border(left=Side(border_style=None, color='FF000000'),
                     right=Side(border_style=None, color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))
        

        for column in range(1,columna_totales+1):
            cell = ws.cell(8,column)
            cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        for column in range(columna_totales+2,columna_totales+4):
            cell = ws.cell(8,column)
            #cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        for column_cells in ws.columns:
            #length = max(len(self.as_text(cell.value)) for cell in column_cells)
            length = len(self.as_text(column_cells[7].value))
            ws.column_dimensions[column_cells[7].column_letter].width = length+5

        ws.column_dimensions['A'].width = 16

        for cell in ws['A']:
            cell.font = Font(bold=True)


        for cell in ws[str(sumas_row)+":"+str(sumas_row)]:
            cell.border = cell_border_sumas
            cell.font = Font(bold=True)


        for i in range(2,sumas_row+1):
            for j in range(2,columna_totales+1):
                ws.cell(i,j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        ws.cell(sumas_row,1).fill = PatternFill(start_color="4ac6ff", end_color="4ac6ff", fill_type = "solid")
        ws.cell(sumas_row,1).border = cell_border

    def style_ws_ingresos(self, ws, columna_totales, sumas_row):
        cell_border = Border(left=Side(border_style=None, color='FF000000'),
                     right=Side(border_style=None, color='FF000000'),
                     top=Side(border_style='medium', color='FF000000'),
                     bottom=Side(border_style='medium', color='FF000000'))

        cell_border_sumas = Border(left=Side(border_style=None, color='FF000000'),
                     right=Side(border_style=None, color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))

        for column in range(1,19):
            cell = ws.cell(8,column)
            cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 9
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 9

        # for cell in ws['A']:
        #     cell.font = Font(bold=True)



        for column in range(1,19):
            cell = ws.cell(sumas_row,column)
            cell.border = cell_border_sumas
            cell.font = Font(bold=True)


        for i in range(9,sumas_row+1):
            for j in range(8,14):
                ws.cell(i,j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        for column in range(20,27):
            cell = ws.cell(10,column)
            cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
            cell = ws.cell(27,column)
            cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        for column in range(20,27):
            cell = ws.cell(23,column)
            cell.border = cell_border_sumas
            cell.font = Font(bold=True)
            cell = ws.cell(40,column)
            cell.border = cell_border_sumas
            cell.font = Font(bold=True)

        ws.column_dimensions['T'].width = 12
        ws.column_dimensions['U'].width = 12
        ws.column_dimensions['V'].width = 12
        ws.column_dimensions['W'].width = 12
        ws.column_dimensions['X'].width = 12
        ws.column_dimensions['Y'].width = 12
        ws.column_dimensions['Z'].width = 12



    def calculaAgregados(self, df, ws_cats, variable):
        # for col in df.columns:
        #     print(col)
        print(df.head())
        self.agregaMembrete(ws_cats, 4)
        ws_cats.cell(7,1).value = " "
        
        por_categorias = df.groupby(['mes', 'tipo'], as_index=False).agg({variable:sum})
        por_categorias_wide = por_categorias.pivot_table(index="mes",columns=['tipo'],values=variable,fill_value= 0)
        por_categorias_wide.reset_index(inplace=True)
        por_categorias_wide['mes'] = pd.Categorical(por_categorias_wide['mes'], self.todos_los_meses)
        por_categorias_wide = por_categorias_wide.sort_values("mes")
        por_categorias_wide['mes'] = por_categorias_wide.mes.astype(str)
        por_categorias_wide_no_deducibles = por_categorias_wide.loc[:,por_categorias_wide.columns.isin(['Gasto personal', 'No deducible',])]
        por_categorias_wide_deducibles = por_categorias_wide.drop(['Gasto personal',  'Nómina', 'Nomina', 'No deducible', 'Pago nomina', 'Pago nómina'], axis=1, errors='ignore')
        #
        for r in dataframe_to_rows(por_categorias_wide_deducibles, index=False, header=True):
            ws_cats.append(r)
        #print(por_categorias_wide)

        if variable == "subTotal":
            col_sum = "I"

        if variable == "impuestos":
            col_sum = "J"



        self.numeroDeColumnas = len(por_categorias_wide_deducibles.columns)
        self.columna_totales = self.numeroDeColumnas + 1
        self.sumas_row = len(por_categorias_wide.index)+9
        ws_cats.cell(8,self.columna_totales, "Total")
        ws_cats.cell(self.sumas_row,1,"Anual")

        for i in range(2,self.columna_totales):
            letra = get_column_letter(i)
            for j in range(9,self.sumas_row):
                ws_cats.cell(j,i,"=SUMIFS(Conceptos!"+col_sum+":"+col_sum+",Conceptos!L:L,"+letra+"8,Conceptos!A:A,A"+str(j)+',Conceptos!M:M,"Pagado")')


         
        ws_cats.cell(8,self.columna_totales+2,'Gasto personal')
        ws_cats.cell(8,self.columna_totales+3,'No deducible')
        for i in range(self.columna_totales+2,self.columna_totales+4):
            letra = get_column_letter(i)
            for j in range(9,self.sumas_row):
                ws_cats.cell(j,i,"=SUMIFS(Conceptos!"+col_sum+":"+col_sum+",Conceptos!L:L,"+letra+"8,Conceptos!A:A,A"+str(j)+',Conceptos!M:M,"Pagado")')

        for i in range(2,self.columna_totales):
            letra = get_column_letter(i)
            ws_cats.cell(self.sumas_row,i,"=SUM("+letra+ "9:"+letra+ str(self.sumas_row-1)+")")

        letra_final = get_column_letter(self.numeroDeColumnas)
        for i in range(9,self.sumas_row):
            ws_cats.cell(i,self.columna_totales,"=SUM(B"+str(i)+ ":"+letra_final+ str(i)+")")

        letra_sumas = get_column_letter(self.columna_totales)
        letraGastosP = get_column_letter(self.columna_totales+2)
        letraNoDeducibles = get_column_letter(self.columna_totales+3)
        ws_cats.cell(self.sumas_row,self.columna_totales,"=SUM("+letra_sumas+"9:"+letra_sumas+str(self.sumas_row-1)  +")")
        ws_cats.cell(self.sumas_row,self.columna_totales+2,"=SUM("+letraGastosP+"9:"+letraGastosP+str(self.sumas_row-1)  +")")
        ws_cats.cell(self.sumas_row,self.columna_totales+3,"=SUM("+letraNoDeducibles+"9:"+letraNoDeducibles+str(self.sumas_row-1)  +")")


    def cuadroRegimen(self, ws_regimenes, row_inicial, column_inicial, regimen_key):
        cell_border = Border(left=Side(border_style='medium', color='FF000000'),
                     right=Side(border_style='medium', color='FF000000'),
                     top=Side(border_style='medium', color='FF000000'),
                     bottom=Side(border_style='medium', color='FF000000'))

        cell_border_sumas = Border(left=Side(border_style=None, color='FF000000'),
                     right=Side(border_style=None, color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))

        letra_meses = get_column_letter(column_inicial)
        letra_total = get_column_letter(column_inicial+6)
        ws_regimenes.merge_cells(letra_meses+str(row_inicial)+":"+letra_total+str(row_inicial))
        cell = ws_regimenes.cell(row=row_inicial, column=column_inicial)
        cell.value = self.regimen[regimen_key] + " (Régimen Fiscal: " + regimen_key + ")"
        # alignment = copy(cell.alignment)
        # alignment.wrapText=True
        # cell.alignment = alignment
        cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)   
        cell.border = cell_border        
        ws_regimenes.row_dimensions[row_inicial].height = 30
        ws_regimenes.cell(row_inicial+1, column_inicial, "Mes")
        ws_regimenes.cell(row_inicial+1, column_inicial+1, "SUBTOTAL")
        ws_regimenes.cell(row_inicial+1, column_inicial+2, "I.V.A.")
        ws_regimenes.cell(row_inicial+1, column_inicial+3, "IMPORTE")
        ws_regimenes.cell(row_inicial+1, column_inicial+4, "RET ISR")
        ws_regimenes.cell(row_inicial+1, column_inicial+5, "RET IVA")
        ws_regimenes.cell(row_inicial+1, column_inicial+6, "T O T A L")

        ws_regimenes.cell(row_inicial+2, column_inicial, "ENERO")
        ws_regimenes.cell(row_inicial+3, column_inicial, "FEBRERO")
        ws_regimenes.cell(row_inicial+4, column_inicial, "MARZO")
        ws_regimenes.cell(row_inicial+5, column_inicial, "ABRIL")
        ws_regimenes.cell(row_inicial+6, column_inicial, "MAYO")
        ws_regimenes.cell(row_inicial+7, column_inicial, "JUNIO")
        ws_regimenes.cell(row_inicial+8, column_inicial, "JULIO")
        ws_regimenes.cell(row_inicial+9, column_inicial, "AGOSTO")
        ws_regimenes.cell(row_inicial+10, column_inicial, "SEPTIEMBRE")
        ws_regimenes.cell(row_inicial+11, column_inicial, "OCTUBRE")
        ws_regimenes.cell(row_inicial+12, column_inicial, "NOVIEMBRE")
        ws_regimenes.cell(row_inicial+13, column_inicial, "DICIEMBRE")
        letra_meses = get_column_letter(column_inicial)
        for renglonMes in range(row_inicial+2,row_inicial+14):
            celda = ws_regimenes.cell(renglonMes, column_inicial+1)
            celda.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            celda.value = '=SUMIFS(Ingresos!H:H,Ingresos!B:B,'+letra_meses+str(renglonMes)+',Ingresos!O:O,"Pagado",Ingresos!R:R,"'+regimen_key+'")'
            celda = ws_regimenes.cell(renglonMes, column_inicial+2)
            celda.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            celda.value = '=SUMIFS(Ingresos!I:I,Ingresos!B:B,'+letra_meses+str(renglonMes)+',Ingresos!O:O,"Pagado",Ingresos!R:R,"'+regimen_key+'")'
            celda = ws_regimenes.cell(renglonMes, column_inicial+3)
            celda.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            celda.value = '=SUMIFS(Ingresos!J:J,Ingresos!B:B,'+letra_meses+str(renglonMes)+',Ingresos!O:O,"Pagado",Ingresos!R:R,"'+regimen_key+'")'
            celda = ws_regimenes.cell(renglonMes, column_inicial+4)
            celda.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            celda.value = '=SUMIFS(Ingresos!K:K,Ingresos!B:B,'+letra_meses+str(renglonMes)+',Ingresos!O:O,"Pagado",Ingresos!R:R,"'+regimen_key+'")'
            celda = ws_regimenes.cell(renglonMes, column_inicial+5)
            celda.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            celda.value = '=SUMIFS(Ingresos!L:L,Ingresos!B:B,'+letra_meses+str(renglonMes)+',Ingresos!O:O,"Pagado",Ingresos!R:R,"'+regimen_key+'")'
            celda = ws_regimenes.cell(renglonMes, column_inicial+6)
            celda.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            celda.value = '=SUMIFS(Ingresos!M:M,Ingresos!B:B,'+letra_meses+str(renglonMes)+',Ingresos!O:O,"Pagado",Ingresos!R:R,"'+regimen_key+'")'


        #letra_sumas = get_column_letter(self.columna_totales)
        letra_subtotal = get_column_letter(column_inicial+1)
        letra_iva = get_column_letter(column_inicial+2)
        letra_importe = get_column_letter(column_inicial+3)
        letra_retisr = get_column_letter(column_inicial+4)
        letra_retiva = get_column_letter(column_inicial+5)
        letra_total = get_column_letter(column_inicial+6)

        ws_regimenes.cell(row_inicial+14, column_inicial+1, "=SUM("+letra_subtotal+str(row_inicial+2)+":"+letra_subtotal+str(row_inicial+13)+")")
        ws_regimenes.cell(row_inicial+14, column_inicial+2, "=SUM("+letra_iva+str(row_inicial+2)+":"+letra_iva+str(row_inicial+13)+")")
        ws_regimenes.cell(row_inicial+14, column_inicial+3, "=SUM("+letra_importe+str(row_inicial+2)+":"+letra_importe+str(row_inicial+13)+")")
        ws_regimenes.cell(row_inicial+14, column_inicial+4, "=SUM("+letra_retisr+str(row_inicial+2)+":"+letra_retisr+str(row_inicial+13)+")")
        ws_regimenes.cell(row_inicial+14, column_inicial+5, "=SUM("+letra_retiva+str(row_inicial+2)+":"+letra_retiva+str(row_inicial+13)+")")
        ws_regimenes.cell(row_inicial+14, column_inicial+6, "=SUM("+letra_total+str(row_inicial+2)+":"+letra_total+str(row_inicial+13)+")")


        for column in range(column_inicial,column_inicial+7):
            cell = ws_regimenes.cell(row_inicial+1,column)
            cell.fill = PatternFill(start_color="8ccbff", end_color="8ccbff", fill_type = "solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
            

        for column in range(column_inicial,column_inicial+7):
            cell = ws_regimenes.cell(row_inicial+14,column)
            cell.border = cell_border_sumas
            cell.font = Font(bold=True)
            




    def hazWsRegimenes(self):
        workbook = load_workbook(self.annual_xlsx_path)
        if not "Regimenes" in workbook.sheetnames:
            ws_regimenes = workbook.create_sheet("Regimenes")
        else:
            ws_regimenes = workbook["Regimenes"]

        self.agregaMembrete(ws_regimenes, 7)
        ws_regimenes.column_dimensions['A'].width = 14
        ws_regimenes.column_dimensions['B'].width = 10
        ws_ingresos = workbook["Ingresos"]
        regimenes = []
        for row_index in range(9,ws_ingresos.max_row):
            row = [cell.value for cell in ws_ingresos[row_index]]
            regimen = row[17]
            if regimen not in regimenes and regimen != None:
                regimenes.append(regimen)
        print(regimenes)
        renglon_inicial = 9
        for regimen in regimenes:
            self.cuadroRegimen(ws_regimenes, renglon_inicial, 1, regimen)
            renglon_inicial+=17

        workbook.save(self.annual_xlsx_path)


    def hazTabDeIngresos(self,paths):## hazWSIngresos?
        if len(self.listaDeFacturasIngresos) > 0:
            workbook = load_workbook(self.annual_xlsx_path)
            if not "Ingresos" in workbook.sheetnames:
                ws_ingresos = workbook.create_sheet("Ingresos")
            else:
                ws_ingresos = workbook["Ingresos"]

            if ws_ingresos.max_row == 1:
                self.agregaMembrete(ws_ingresos, 4)
                ws_ingresos.cell(8, 1, "MesEmision")
                ws_ingresos.cell(8, 2,     "MesPago")
                ws_ingresos.cell(8, 3,     "uuid")
                ws_ingresos.cell(8, 4,     "FECHA")
                ws_ingresos.cell(8, 5,     "RFC (Receptor)")
                ws_ingresos.cell(8, 6,     "RAZON SOCIAL")
                ws_ingresos.cell(8, 7,     "DESCRIPCION")
                ws_ingresos.cell(8, 8,     "SUBTOTAL")
                ws_ingresos.cell(8, 9,     "I.V.A.")
                ws_ingresos.cell(8, 10,     "IMPORTE")
                ws_ingresos.cell(8, 11,     "RET ISR")
                ws_ingresos.cell(8, 12,     "RET IVA")
                ws_ingresos.cell(8, 13,     "T O T A L")
                ws_ingresos.cell(8, 14,     "M-Pago")
                ws_ingresos.cell(8, 15,     "Status")
                ws_ingresos.cell(8, 16,     "complementosDePago")
                ws_ingresos.cell(8, 17,     "Tipo")
                ws_ingresos.cell(8, 18,     "Regimen")
                row = 8
                c = ws_ingresos['A9']
                ws_ingresos.freeze_panes = c
            else:
                max_row_for_a = max((c.row for c in ws_ingresos['C'] if c.value is not None))
                row = max_row_for_a
                #row = ws_ingresos.max_row
            dv = DataValidation(type="list", formula1='"Pendiente,Pagado"', allowBlank=True)
            ws_ingresos.add_data_validation(dv)
            dv_mes = DataValidation(type="list", formula1='"ENERO,FEBRERO,MARZO,ABRIL,MAYO,JUNIO,JULIO,AGOSTO,SEPTIEMBRE,OCTUBRE,NOVIEMBRE,DICIEMBRE,--"', allow_blank=True)
            ws_ingresos.add_data_validation(dv_mes)

            for factura in self.listaDeFacturasIngresos:
                numeroDeMes = int(factura.fechaTimbrado.split("-")[1])
                este_mes = self.todos_los_meses[numeroDeMes-1]
                print(self.yaEstaba)
                if not self.yaEstaba[este_mes]:
                    row += 1
                    
                    dv_mes.add(ws_ingresos.cell(row, 1))
                    dv_mes.add(ws_ingresos.cell(row, 2))
                    ws_ingresos.cell(row, 1, self.todos_los_meses[numeroDeMes-1])
                    if factura.metodoDePago == "PUE" or factura.tipoDeComprobante == "P":
                        ws_ingresos.cell(row, 2, self.todos_los_meses[numeroDeMes-1])
                    if factura.UUID in self.complementosDePago:
                        numeroDeMesP = int(self.complementosDePago[factura.UUID]["fechaUltimoPago"].split("-")[1])
                        ws_ingresos.cell(row, 2, self.todos_los_meses[numeroDeMesP-1])
                    ws_ingresos.cell(row, 3, factura.UUID)
                    ws_ingresos.cell(row, 4, factura.fechaTimbrado)
                    ws_ingresos.cell(row, 5, factura.ReceptorRFC)
                    ws_ingresos.cell(row, 6, factura.ReceptorNombre)
                    ws_ingresos.cell(row, 7, factura.conceptos[0]['descripcion'])
                    ws_ingresos.cell(row, 8, factura.subTotal)
                    ws_ingresos.cell(row, 9, factura.traslados["IVA"]["importe"])
                    ws_ingresos.cell(row, 10, factura.importe)
                    if factura.tipoDeComprobante == "N":
                        ws_ingresos.cell(row, 11, factura.RetencionesISRNomina)
                    else:
                        ws_ingresos.cell(row, 11, factura.retenciones["ISR"])
                    ws_ingresos.cell(row, 12, factura.retenciones["IVA"])
                    ws_ingresos.cell(row, 13, factura.total)
                    ws_ingresos.cell(row, 14, factura.metodoDePago)

                    status = "Pendiente"
                    if factura.metodoDePago == "PUE":
                        status = "Pagado"
                    if factura.metodoDePago == "PPD":
                        if factura.UUID in self.complementosDePago:
                            if factura.total - self.complementosDePago[factura.UUID]["suma"] < 0.5:
                                status = "Pagado"
                    if factura.tipoDeComprobante == "P":
                        status = "Pagado"

                    dv.add(ws_ingresos.cell(row, 15))
                    ws_ingresos.cell(row, 15, status)

                    if factura.UUID in self.complementosDePago:
                        ws_ingresos.cell(row, 16, self.complementosDePago[factura.UUID]["suma"])

                    if factura.tipoDeComprobante == "N":
                        ws_ingresos.cell(row, 17, "Nómina")
                    else:
                        ws_ingresos.cell(row, 17, "Facturado")

                    ws_ingresos.cell(row, 18, factura.EmisorRegimen)

            ws_ingresos.cell(row+1, 8, "=SUM(H2:H"+str(row)+")")
            ws_ingresos.cell(row+1, 9, "=SUM(I2:I"+str(row)+")")
            ws_ingresos.cell(row+1, 10, "=SUM(J2:J"+str(row)+")")
            ws_ingresos.cell(row+1, 11, "=SUM(K2:K"+str(row)+")")
            ws_ingresos.cell(row+1, 12, "=SUM(L2:L"+str(row)+")")
            ws_ingresos.cell(row+1, 13, "=SUM(M2:M"+str(row)+")")

            ws_ingresos.cell(9, 23, "Facturado")#bajo protesta
            ws_ingresos.cell(10, 20, "Mes")
            ws_ingresos.cell(10, 21, "SUBTOTAL")
            ws_ingresos.cell(10, 22, "I.V.A.")
            ws_ingresos.cell(10, 23, "IMPORTE")
            ws_ingresos.cell(10, 24, "RET ISR")
            ws_ingresos.cell(10, 25, "RET IVA")
            ws_ingresos.cell(10, 26, "T O T A L")

            ws_ingresos.cell(11, 20, "ENERO")
            ws_ingresos.cell(12, 20, "FEBRERO")
            ws_ingresos.cell(13, 20, "MARZO")
            ws_ingresos.cell(14, 20, "ABRIL")
            ws_ingresos.cell(15, 20, "MAYO")
            ws_ingresos.cell(16, 20, "JUNIO")
            ws_ingresos.cell(17, 20, "JULIO")
            ws_ingresos.cell(18, 20, "AGOSTO")
            ws_ingresos.cell(19, 20, "SEPTIEMBRE")
            ws_ingresos.cell(20, 20, "OCTUBRE")
            ws_ingresos.cell(21, 20, "NOVIEMBRE")
            ws_ingresos.cell(22, 20, "DICIEMBRE")

            for renglonMes in range(11,22):
                ws_ingresos.cell(renglonMes, 21, '=SUMIFS(H:H,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Facturado")')
                ws_ingresos.cell(renglonMes, 22, '=SUMIFS(I:I,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Facturado")')
                ws_ingresos.cell(renglonMes, 23, '=SUMIFS(J:J,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Facturado")')
                ws_ingresos.cell(renglonMes, 24, '=SUMIFS(K:K,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Facturado")')
                ws_ingresos.cell(renglonMes, 25, '=SUMIFS(L:L,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Facturado")')
                ws_ingresos.cell(renglonMes, 26, '=SUMIFS(M:M,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Facturado")')

            ws_ingresos.cell(23, 21, "=SUM(U11:U22)")
            ws_ingresos.cell(23, 22, "=SUM(V11:V22)")
            ws_ingresos.cell(23, 23, "=SUM(W11:W22)")
            ws_ingresos.cell(23, 24, "=SUM(X11:X22)")
            ws_ingresos.cell(23, 25, "=SUM(Y11:Y22)")
            ws_ingresos.cell(23, 26, "=SUM(Z11:Z22)")


            ws_ingresos.cell(26, 23, "Nómina")
            ws_ingresos.cell(27, 20, "Mes")
            ws_ingresos.cell(27, 21, "SUBTOTAL")
            ws_ingresos.cell(27, 22, "I.V.A.")
            ws_ingresos.cell(27, 23, "IMPORTE")
            ws_ingresos.cell(27, 24, "RET ISR")
            ws_ingresos.cell(27, 25, "RET IVA")
            ws_ingresos.cell(27, 26, "T O T A L")

            ws_ingresos.cell(28, 20, "ENERO")
            ws_ingresos.cell(29, 20, "FEBRERO")
            ws_ingresos.cell(30, 20, "MARZO")
            ws_ingresos.cell(31, 20, "ABRIL")
            ws_ingresos.cell(32, 20, "MAYO")
            ws_ingresos.cell(33, 20, "JUNIO")
            ws_ingresos.cell(34, 20, "JULIO")
            ws_ingresos.cell(35, 20, "AGOSTO")
            ws_ingresos.cell(36, 20, "SEPTIEMBRE")
            ws_ingresos.cell(37, 20, "OCTUBRE")
            ws_ingresos.cell(38, 20, "NOVIEMBRE")
            ws_ingresos.cell(39, 20, "DICIEMBRE")

            for renglonMes in range(28,40):
                ws_ingresos.cell(renglonMes, 21, '=SUMIFS(H:H,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Nómina")')
                ws_ingresos.cell(renglonMes, 22, '=SUMIFS(I:I,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Nómina")')
                ws_ingresos.cell(renglonMes, 23, '=SUMIFS(J:J,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Nómina")')
                ws_ingresos.cell(renglonMes, 24, '=SUMIFS(K:K,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Nómina")')
                ws_ingresos.cell(renglonMes, 25, '=SUMIFS(L:L,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Nómina")')
                ws_ingresos.cell(renglonMes, 26, '=SUMIFS(M:M,B:B,T'+str(renglonMes)+',O:O,"Pagado",Q:Q,"Nómina")')

            ws_ingresos.cell(40, 21, "=SUM(U28:U39)")
            ws_ingresos.cell(40, 22, "=SUM(V28:V39)")
            ws_ingresos.cell(40, 23, "=SUM(W28:W39)")
            ws_ingresos.cell(40, 24, "=SUM(X28:X39)")
            ws_ingresos.cell(40, 25, "=SUM(Y28:Y39)")
            ws_ingresos.cell(40, 26, "=SUM(Z28:Z39)")

            self.style_ws_ingresos(ws_ingresos,17,row+1)

            workbook.save(self.annual_xlsx_path)

    def getDescription(self, clave):
        desc = ""
        try:
            desc = self.concepto[clave]
        except:
            desc = ""
        return desc

    def hazAgregados(self, paths):
        print(self.complementosDePago)
        workbook = load_workbook(self.annual_xlsx_path)
        if not "Conceptos" in workbook.sheetnames:
            ws_todos = workbook.create_sheet("Conceptos")
        else:
            ws_todos = workbook["Conceptos"]

        if "IVA_mensual" in workbook.sheetnames:
            sheet1 = workbook["IVA_mensual"]
            workbook.remove(sheet1)
        if "Egresos_mensual" in workbook.sheetnames:
            sheet1 = workbook["Egresos_mensual"]
            workbook.remove(sheet1)
        if "Categorias" in workbook.sheetnames:
            sheet1 = workbook["Categorias"]
            workbook.remove(sheet1)

        workbook.save(self.annual_xlsx_path)


        ws_cats = workbook.create_sheet("IVA_mensual")
        ws_egresos_mensual = workbook.create_sheet("Egresos_mensual")
        ws_lista_cats = workbook.create_sheet("Categorias")
        r = 0
        for cat in self.lista_categorias_default:
            r += 1
            ws_lista_cats.cell(r, 1, cat)

        self.status_column = 0
        primer_mes = workbook[self.meses[0]]

        for i in range(1,primer_mes.max_column+1):
            if primer_mes.cell(8, i).value == "Status":
                self.status_column = i - 2
        for mes in self.meses:
            ws_mes = workbook[mes]
            for row in range(9,len(ws_mes["A"])): 
                if ws_mes.cell(row, 12).value == "PPD":#11 H
                    print("ajustaria"+ ws_mes.cell(row, 3).value)
                    if ws_mes.cell(row, 3).value in self.complementosDePago:
                        ws_mes.cell(row, 16, self.complementosDePago[ws_mes.cell(row, 3).value]["suma"])
                        if ws_mes.cell(row, 10).value - self.complementosDePago[ws_mes.cell(row, 3).value]["suma"] < 0.5:
                            ws_mes.cell(row, 14, "Pagado")


        print("ws_todos.max_row................................................................",str(ws_todos.max_row))
        if ws_todos.max_row == 1:
            self.agregaMembrete(ws_todos, 4)
            ws_todos.cell(8, 1, "mes")
            ws_todos.cell(8, 2, 'clave_concepto')
            ws_todos.cell(8, 3, 'concepto_sat')
            #self.concepto
            ws_todos.cell(8, 4, 'UUID')
            ws_todos.cell(8, 5, 'cantidad')
            ws_todos.cell(8, 6, 'descripcion')
            ws_todos.cell(8, 7, 'importeConcepto')
            ws_todos.cell(8, 8, 'descuento')
            ws_todos.cell(8, 9, 'subTotal')
            ws_todos.cell(8, 10, 'impuestos')
            ws_todos.cell(8, 11, 'total')
            ws_todos.cell(8, 12, 'tipo')
            ws_todos.cell(8, 13, 'status')
            ws_todos.column_dimensions['A'].width = 9
            ws_todos.column_dimensions['B'].width = 10
            ws_todos.column_dimensions['C'].width = 40
            ws_todos.column_dimensions['D'].width = 40
            ws_todos.column_dimensions['E'].width = 10
            ws_todos.column_dimensions['F'].width = 40
            ws_todos.column_dimensions['G'].width = 9
            ws_todos.column_dimensions['H'].width = 9
            ws_todos.column_dimensions['I'].width = 9
            ws_todos.column_dimensions['J'].width = 9
            ws_todos.column_dimensions['K'].width = 9
            ws_todos.column_dimensions['L'].width = 15
            ws_todos.column_dimensions['M'].width = 9
            row = 8
        else:
            row = ws_todos.max_row

        # for mes in meses:
        #     if mes in meses_folders:
        #         for concepto in self.conceptos[mes]:

        #dv_categorias = DataValidation(type="list", formula1='"{}"'.format(self.texto_para_validacion), allow_blank=True)
        dv_categorias = DataValidation(type="list", formula1="=Categorias!A$1:A$"+str(len(self.lista_categorias_default)), allow_blank=True)
        c = ws_mes['A9']
        ws_todos.freeze_panes = c
        ws_todos.add_data_validation(dv_categorias)
        for concepto in self.conceptos:
            if not self.yaEstaba[concepto['mes']]:
                row += 1
                clave = concepto['clave_concepto']
                ws_todos.cell(row, 1, concepto['mes'])
                ws_todos.cell(row, 2, clave)
                ws_todos.cell(row, 3, self.getDescription(clave))
                ws_todos.cell(row, 4, concepto['UUID'])
                ws_todos.cell(row, 5, concepto['cantidad'])
                ws_todos.cell(row, 6, concepto['descripcion'])
                ws_todos.cell(row, 7, concepto['importeConcepto'])
                ws_todos.cell(row, 8, concepto['descuento'])
                ws_todos.cell(row, 9, concepto['importeConcepto'] - concepto['descuento'])
                ws_todos.cell(row, 10, concepto['impuestos'])
                ws_todos.cell(row, 11, (concepto['importeConcepto'] - concepto['descuento']) + concepto['impuestos'])
                dv_categorias.add(ws_todos.cell(row, 12))
                ws_todos.cell(row, 12, concepto['tipo'])
                ws_todos.cell(row, 13, "=VLOOKUP(D"+str(row)+","+concepto['mes']+"!C:Q,"+str(self.status_column)+",FALSE)")

        df = pd.DataFrame(self.conceptos)



        self.calculaAgregados(df, ws_cats, 'impuestos')
        self.style_ws(ws_cats, self.columna_totales, self.sumas_row)

        self.calculaAgregados(df, ws_egresos_mensual, 'subTotal')
        self.style_ws(ws_egresos_mensual, self.columna_totales, self.sumas_row)


        workbook.save(self.annual_xlsx_path)

    def agregaMembrete(self, ws, columna):
        for column in range(1,28):
            for row in range(1,8):
                cell = ws.cell(row,column)
                cell.fill = PatternFill(start_color="8D99AD", end_color="97cffc", fill_type = "solid")
                cell.font = Font(bold=True)
                
        img = openpyxl.drawing.image.Image(join(scriptDirectory,'logo_s.png'))
        img.anchor = 'B2'
        ws.add_image(img)
        ws.cell(2, columna, "Nombre: ")
        ws.cell(3, columna, "RFC: ")
        ws.cell(2, columna+1, self.nombre)
        ws.cell(3, columna+1, self.rfc)
        ws.cell(4, columna, "Contador: ")
        ws.cell(5, columna, "Fecha de Actualización: ") 
        ws.cell(5, columna+1, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        letra = get_column_letter(columna)
        for row in ws[letra + "2":letra + "6"]:
            for cell in row:
                print(cell.value)
                cell.alignment = Alignment(horizontal="right")

        img = openpyxl.drawing.image.Image(join(scriptDirectory,'logo_excel.png'))
        img.anchor = 'M2'
        ws.add_image(img)

        


    def agregaMes(self, mes):
        if os.path.isfile(self.annual_xlsx_path):
            workbook = load_workbook(self.annual_xlsx_path)
            if not mes in workbook.sheetnames:
                self.yaEstaba[mes] = False
                ws_mes = workbook.create_sheet(mes, 0)
            else:
                self.yaEstaba[mes] = True
        else:
            workbook = Workbook()
            ws_mes = workbook.create_sheet(mes)
            sheet1_name = workbook.get_sheet_names()[0]
            sheet1 = workbook[sheet1_name]
            workbook.remove_sheet(sheet1)
            self.yaEstaba[mes] = False

        #TUA	IEPS	ISH

        if not self.yaEstaba[mes]:
            self.agregaMembrete(ws_mes, 4)


            ws_mes.cell(8, 1, "clave_ps")
            ws_mes.cell(8, 2,     "Fecha")
            ws_mes.cell(8, 3,     "UUID")
            ws_mes.cell(8, 4,     "Nombre")
            ws_mes.cell(8, 5,     "RFC")
            ws_mes.cell(8, 6,     "Concepto")
            ws_mes.cell(8, 7,     "Sub")
            ws_mes.cell(8, 8,     "Descuento")
            ws_mes.cell(8, 9,     "IVA")
            ws_mes.cell(8, 10,     "TUA")
            ws_mes.cell(8, 11,     "ISH")
            ws_mes.cell(8, 12,     "IEPS")
            ws_mes.cell(8, 13,     "Total")
            ws_mes.cell(8, 14,     "F-Pago")
            ws_mes.cell(8, 15,     "M-Pago")
            ws_mes.cell(8, 16,     "Tipo")
            ws_mes.cell(8, 17,     "Status")
            ws_mes.cell(8, 18,     "TipoDeComprobante")
            ws_mes.cell(8, 19,     "complementosDePago")
            ws_mes.cell(8, 20,     "Folio")
            ws_mes.cell(8, 21,     "Serie")

            c = ws_mes['A9']
            ws_mes.freeze_panes = c

        

            ws_mes.column_dimensions['A'].width = 10
            ws_mes.column_dimensions['B'].width = 20
            ws_mes.column_dimensions['C'].width = 40
            ws_mes.column_dimensions['D'].width = 35
            ws_mes.column_dimensions['E'].width = 16
            ws_mes.column_dimensions['F'].width = 30
            ws_mes.column_dimensions['G'].width = 9
            ws_mes.column_dimensions['H'].width = 9
            ws_mes.column_dimensions['I'].width = 9
            ws_mes.column_dimensions['J'].width = 9
            ws_mes.column_dimensions['K'].width = 9
            ws_mes.column_dimensions['L'].width = 9
            ws_mes.column_dimensions['M'].width = 9
            ws_mes.column_dimensions['N'].width = 20
            ws_mes.column_dimensions['O'].width = 7
            ws_mes.column_dimensions['P'].width = 20
            ws_mes.column_dimensions['Q'].width = 20
            ws_mes.column_dimensions['R'].width = 17
            ws_mes.column_dimensions['S'].width = 17
            ws_mes.column_dimensions['T'].width = 20
            ws_mes.column_dimensions['U'].width = 20

            dv = DataValidation(type="list", formula1='"Pendiente,Pagado"', allow_blank=True)
            ws_mes.add_data_validation(dv)

            bg = PatternFill(bgColor = "faf3c5")
            style = DifferentialStyle(fill=bg) 
            rule = Rule(type="expression", dxf=style)
            rule.formula = ['$Q9="Pendiente"']

            row = 8
            for factura in self.facturas[self.mes]:
                row += 1
                ws_mes.cell(row, 1, factura.conceptos[0]['clave_concepto'])
                ws_mes.cell(row, 2, factura.fechaTimbrado)
                ws_mes.cell(row, 3, factura.UUID)
                ws_mes.cell(row, 4, factura.EmisorNombre)
                ws_mes.cell(row, 5, factura.EmisorRFC)
                ws_mes.cell(row, 6, factura.conceptos[0]['descripcion'])
                if factura.tipoDeComprobante == "E":
                    ws_mes.cell(row, 7, 0.0 - factura.subTotal)
                    ws_mes.cell(row, 8, 0.0 - factura.descuento)
                    ws_mes.cell(row, 9, 0.0 - factura.traslados["IVA"]["importe"])
                    ws_mes.cell(row, 10, 0.0)
                    ws_mes.cell(row, 11, 0.0)
                    ws_mes.cell(row, 12, 0.0)
                    ws_mes.cell(row, 13, 0.0 - factura.total)
                else:
                    ws_mes.cell(row, 7, factura.subTotal)
                    ws_mes.cell(row, 8, factura.descuento)
                    ws_mes.cell(row, 9, factura.traslados["IVA"]["importe"])
                    ws_mes.cell(row, 10, factura.trasladosLocales["TUA"]["importe"])
                    ws_mes.cell(row, 11, factura.trasladosLocales["ISH"]["importe"])
                    ws_mes.cell(row, 12, factura.traslados["IEPS"]["importe"])
                    ws_mes.cell(row, 13, factura.total)
                ws_mes.cell(row, 14, factura.formaDePagoStr)
                ws_mes.cell(row, 15, factura.metodoDePago)
                ws_mes.cell(row, 16, factura.conceptos[0]['tipo'])
                status = "Pendiente"
                if factura.metodoDePago == "PUE":
                    status = "Pagado"
                if factura.metodoDePago == "PPD":
                    if factura.UUID in self.complementosDePago:
                        if factura.total - self.complementosDePago[factura.UUID]["suma"] < 0.5:
                            status = "Pagado"
                if factura.tipoDeComprobante == "P":
                    status = "Pagado"

                dv.add(ws_mes.cell(row, 17))
                ws_mes.cell(row, 17, status)
                ws_mes.cell(row, 18, factura.tipoDeComprobante)
                if factura.UUID in self.complementosDePago:
                    ws_mes.cell(row, 19, self.complementosDePago[factura.UUID]["suma"])

                ws_mes.cell(row, 20, factura.folio)
                ws_mes.cell(row, 21, factura.serie)

                if factura.tipoDeComprobante == "P":
                    print("segun "+ factura.UUID + "del mes " +mes+ ", aqui buscaria en todos los meses el uuid "+factura.IdDocumento+" y si encuentra su factura modificaria, la columna 13 del renglon de esa factura en el mes que esté, a Pagado")
            
            FullRange = "A8:" + get_column_letter(ws_mes.max_column) + str(ws_mes.max_row)
            ws_mes.auto_filter.ref = FullRange
            #aqui haria el renglón de hasta abajo de sunmas?
            if row > 9:
                ws_mes.conditional_formatting.add("A9:U"+str(row), rule)
                ws_mes.cell(row + 1, 6, "Suma de Pagadas")
                for columna_suma in range(7,14):
                    letra = get_column_letter(columna_suma)
                    ws_mes.cell(row + 1, columna_suma, "=SUMIFS(" + letra + "9:" + letra + str(row) +",Q9:Q" + str(row) + ',"Pagado"' + ")")
                cell_border_sumas = Border(left=Side(border_style=None, color='FF000000'),
                                           right=Side(border_style=None, color='FF000000'),
                                           top=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='thin', color='FF000000'))
                for cell in ws_mes[str(row + 1)+":"+str(row + 1)]:
                    cell.border = cell_border_sumas
                    cell.font = Font(bold=True)
            
            workbook.save(self.annual_xlsx_path)

    def hazResumenDiot(self,currentDir):
        workbook = load_workbook(os.path.join(appDataDirectory,"template_diot.xlsx"))
        ws_rfc = workbook[workbook.get_sheet_names()[0]]
        cliente = os.path.split(self.cliente_path)[1]
        xlsx_path = os.path.join(currentDir,os.path.join("huiini","resumen_" + cliente + "_" + self.mes + ".xlsx"))
        #workbook = xlsxwriter.Workbook(xlsx_path)
        #worksheet = workbook.add_worksheet("por_RFC")
        # workbook = Workbook()
        # ws_rfc = workbook.create_sheet("por_RFC")
        # sheet1_name = workbook.get_sheet_names()[0]
        # sheet1 = workbook[sheet1_name]
        # workbook.remove_sheet(sheet1)

        # ws_rfc.cell(1, 1,     "RFC")
        # ws_rfc.cell(1, 2,     "SUBTOTAL")
        # ws_rfc.cell(1, 3,     "DESCUENTO")
        # ws_rfc.cell(1, 4,     "IMPORTE")
        # ws_rfc.cell(1, 5,     "IVA")
        # ws_rfc.cell(1, 6,     "TOTAL")

        row = 5
        for key, value in self.sumaRFC[self.mes].items():
            row += 1
            if row > 9:
                ws_rfc.insert_rows(row)
                for c in range(1,23):
                    cell = ws_rfc.cell(row + 1, c)
                    new_cell = ws_rfc.cell(row, c)
                    new_cell._style = copy(cell._style)
            ws_rfc.cell(row, 3, key)
            ws_rfc.cell(row, 8, value['importeStr'])
            # ws_rfc.cell(row, 2, value['subTotal'])
            # ws_rfc.cell(row, 3, value['descuento'])
            # ws_rfc.cell(row, 4, value['trasladoIVA'])factura.EmisorNombre
            ws_rfc.cell(row, 5, value['nombre'])
            ##factura.EmisorNombre
            ws_rfc.cell(row, 20, value['trasladoIVAStr'])
            # ws_rfc.cell(row, 6, value['total'])
        ws_rfc.cell(row+2, 8, "=SUM(H6:H"+str(row)+")")
        ws_rfc.cell(row+2, 20, "=SUM(T6:T"+str(row)+")")


        #worksheet2 = workbook.add_worksheet("por_Factura")
        ws_factura = workbook.create_sheet("por_Factura")
        ws_factura.cell(1, 1, "clave_ps")
        ws_factura.cell(1, 2,     "Fecha")
        ws_factura.cell(1, 3,     "UUID")
        ws_factura.cell(1, 4,     "Nombre")
        ws_factura.cell(1, 5,     "RFC")
        ws_factura.cell(1, 6,     "Concepto")
        ws_factura.cell(1, 7,     "Sub")
        ws_factura.cell(1, 8,     "IVA")
        ws_factura.cell(1, 9,     "Total")
        ws_factura.cell(1, 10,     "F-Pago")
        ws_factura.cell(1, 11,     "M-Pago")
        ws_factura.cell(1, 12,     "Tipo")

        row = 1
        for factura in self.facturas[self.mes]:
            row += 1
            ws_factura.cell(row, 1, factura.conceptos[0]['clave_concepto'])
            ws_factura.cell(row, 2, factura.fechaTimbrado)
            ws_factura.cell(row, 3, factura.UUID)
            ws_factura.cell(row, 4, factura.EmisorNombre)
            ws_factura.cell(row, 5, factura.EmisorRFC)
            ws_factura.cell(row, 6, factura.conceptos[0]['descripcion'])
            ws_factura.cell(row, 7, factura.subTotal)
            ws_factura.cell(row, 8, factura.traslados["IVA"]["importe"])
            ws_factura.cell(row, 9, factura.total)
            ws_factura.cell(row, 10, factura.formaDePagoStr)
            ws_factura.cell(row, 11, factura.metodoDePago)
            ws_factura.cell(row, 12, factura.conceptos[0]['tipo'])

        row += 1
        ws_factura.cell(row, 7,     "=SUM(G2:G"+str(row-1)+")")
        ws_factura.cell(row, 8,     "=SUM(H2:H"+str(row-1)+")")
        ws_factura.cell(row, 9,     "=SUM(I2:I"+str(row-1)+")")

        workbook.save(xlsx_path)


    def hazListadeUuids(self):
        self.listadeUuids = []
        for renglon in range(self.numeroDeFacturasValidas[self.mes]):
            self.listadeUuids.append(self.tables[self.mes].item(renglon,1).text())


    def handleHeaderMenu(self, pos):
        menu = QMenu()
        deleteAction = QAction('&Delete', self)
        #deleteAction = QtGui.QAction("Delete")
        deleteAction.triggered.connect(lambda: self.quitaRenglon(self.tables[self.mes].verticalHeader().logicalIndexAt(pos)))
        menu.addAction(deleteAction)

        menu.exec_(QtGui.QCursor.pos())

    def quitaRenglon(self,row):
        elNombre = self.tables[self.mes].item(row,2).text()
        suRFC = ""
        for factura in self.facturas[self.mes]:
            if factura.UUID == elNombre:
                print("i found it!")
                suRFC = factura.EmisorRFC

                break


        suSubtotal = float(self.tables[self.mes].item(row,6).text())
        suDescuento = float(self.tables[self.mes].item(row,7).text())
        suTrasladoIVA = float(self.tables[self.mes].item(row,8).text())
        suImporte = float(self.tables[self.mes].item(row,6).text())-float(self.tables[self.mes].item(row,7).text())
        self.tables[self.mes].removeRow(row)

        if suRFC in self.sumaRFC[self.mes]:
            self.sumaRFC[self.mes][suRFC]['subTotal'] -= suSubtotal
            self.sumaRFC[self.mes][suRFC]['descuento'] -= suDescuento
            self.sumaRFC[self.mes][suRFC]['trasladoIVA'] -= suTrasladoIVA
            self.sumaRFC[self.mes][suRFC]['importe'] -= suImporte

            if math.fabs(self.sumaRFC[self.mes][suRFC]['subTotal']) < 0.0001 and math.fabs(self.sumaRFC[self.mes][suRFC]['descuento']) < 0.0001 and math.fabs(self.sumaRFC[self.mes][suRFC]['trasladoIVA']) < 0.0001 and math.fabs(self.sumaRFC[self.mes][suRFC]['importe']) < 0.0001:
                self.sumaRFC[self.mes].pop(suRFC,0)


        self.numeroDeFacturasValidas[self.mes] -= 1
        self.sumale(1)


        self.hazResumenDiot(self.esteFolder)

    def sumale(self, renglonResumen=0):
        total_col = 13
        
        for i in range(6,self.tables[self.mes].columnCount()):
            estaTabla = self.tables[self.mes]
            if estaTabla.horizontalHeaderItem(i).text() == "Total":
                total_col =  i
        print(str(self.numeroDeFacturasValidas[self.mes]))
        for columna in range(6,total_col+1):
            suma = 0
            for renglon in range(self.numeroDeFacturasValidas[self.mes]):
                try:
                    suma += float(self.tables[self.mes].item(renglon, columna).text().replace(",",""))
                except:
                    print("no puedo sumar")

    def ponEncabezado(self,lista_columnas,tabName):
        n = -1
        for columna in lista_columnas:
            n += 1
            self.tables[tabName].setHorizontalHeaderItem (n, QTableWidgetItem(columna))

    def meDoblePicaronXML(self, row,column):
        print("me picaron en : " +str(row)+", " +str(column))

        tabName = self.tabWidget.tabText(self.tabWidget.currentIndex())
        print(tabName)
        folder_mes = ""
        for file in listdir(self.year_folder):
            if tabName in file:
                folder_mes = join(self.year_folder,file,"EGRESOS")
        if column == 2:
            esteUUID = self.tables[tabName].item(row, 2).text().lower()
            for root, dirs, files in os.walk(folder_mes, topdown=False):
                for name in files:
                    if esteUUID in name.lower() and name.endswith("xml"):
                        xmlpath = os.path.join(root, name)
            try:
                print("este guey me pico:"+xmlpath)
                os.startfile(xmlpath)
            except:
                print("el sistema no tiene una aplicacion por default para abrir xmls")
                print(xmlpath)
                QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir xmls" )

        if column == 0:
            pdf = join(join(folder_mes,"huiini"),self.tables[tabName].item(row, 2).text()+".pdf")
            try:
                print("este guey me pico:"+pdf)
                os.startfile(pdf)
            except:
                print ("el sistema no tiene una aplicacion por default para abrir pdfs")
                QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir pdfs" )


    def meDoblePicaronResumen(self, row,column):
        print("me picaron en : " +str(row)+", " +str(column))
        try:
            print("este guey me pico:"+self.excel_path)
            os.startfile(self.excel_path)
            
        except:
            print ("el sistema no tiene una aplicacion por default para abrir exceles")
            QMessageBox.information(self, "Information", "El sistema no tiene una aplicación por default para abrir exceles" )



    def cambiaImpresora(self):
        # self.tabWidget.setCurrentIndex(1)
        #self.listaDeImpresoras.setEnabled(True)
        self.impresoras = impresoras_widget()
        for (a,b,name,d) in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL):
            self.impresoras.lista.addItem(name)
        self.impresoras.exec()

    def cancelaImpresion(self):
        print("cancelaria")
        phandle = win32print.OpenPrinter(win32print.GetDefaultPrinter())

        print_jobs = win32print.EnumJobs(phandle, 0, -1, 1)
        for job in print_jobs:

            print(job['TotalPages'])

            if(job['TotalPages'] >= 1):
                print(type(job))
                win32print.SetJob(phandle, job['JobId'], 0, None, win32print.JOB_CONTROL_DELETE)

        win32print.ClosePrinter(phandle)



    def imprime(self):
        #objetosMagicosOrdenados = sorted(self.objetosMagicos, key=lambda objetosMagicos: objetosMagicos.fecha)
        self.actionCancelar_Impresi_n.setEnabled(True)
        tabName = self.tabWidget.tabText(self.tabWidget.currentIndex())
        print(tabName)
        folder_mes = ""
        for file in listdir(self.year_folder):
            if tabName in file:
                folder_mes = join(self.year_folder,file,"EGRESOS","huiini")

        # if tabName == "Ingresos":
        #     pdf_path1 = join(self.year_folder,file,"EGRESOS","huiini")
        #     pdf_path2 =

        for renglon in range(0,self.tables[tabName].rowCount()-1):
            uuid = self.tables[tabName].item(renglon,2).text()
            columnaTotal = 0
            for i in range(0,self.tables[tabName].columnCount()-1):
                esteHeaderItem = self.tables[tabName].horizontalHeaderItem(i)
                if esteHeaderItem.text() == "Total":
                    columnaTotal = i

            
            if float(self.tables[tabName].item(renglon,columnaTotal).text().replace(",","")) < 0.01:
                print("no imprimo facturas con total menores a 0")
            else:
                if self.tables[tabName].item(renglon,columnaTotal+2).text() == "None":
                    print("no imprimo pagos seño")
                else:
                    pdf_path = '"'+join(folder_mes, uuid+".pdf")+'"'
                    args = '"'+self.gswin64c_path+'" ' \
                            '-sDEVICE=mswinpr2 ' \
                            '-dBATCH ' \
                            '-dNOPAUSE ' \
                            '-dFitPage ' \
                            '-dQueryUser=3 '
                    ghostscript = args + pdf_path
                    call(ghostscript, shell=True)
        
        #hh = win32api.ShellExecute(0, "print", join(join(self.esteFolder,"huiini"), "resumenDiot.pdf") , None,  ".",  0)
    def esteItem(self, text, tooltip):
        item = QTableWidgetItem(text)
        item.setToolTip(tooltip)
        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
        return item

    def esteCenteredItem(self, text, tooltip):
        item = QTableWidgetItem(QtCore.QLocale().toString(float(text),'f', 2))
        #item.setStyleSheet("padding :15px")
        item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
        QtCore.QLocale().toString(float(text),'f', 2)
        item.setToolTip(tooltip)
        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
        return item


    def hazPDFs(self):
        contador = -1
        pdf_folder = join(self.esteFolder,"huiini")
        for factura in self.facturas[self.mes]:
            contador += 1
            if factura.has_pdf == False:
                xml_name = basename(factura.xml_path)
                factura.conviertemeEnTex()
                factura.conviertemeEnPDF(pdfs_folder = pdf_folder)


                factura.has_pdf = True
                self.tables[self.mes].setCellWidget(contador,0, ImgWidgetPalomita(self))

                #     else:
                #         self.tableWidget_xml.setCellWidget(contador,0, ImgWidgetTache(self))
                # except:
                #     self.tableWidget_xml.setCellWidget(contador,0, ImgWidgetTache(self))
    def borraAuxiliares(self):
        for root, dirs, files in os.walk(self.esteFolder, topdown=False):
            for name in files:
                if name.endswith(".tex") or name.endswith(".log") or name.endswith(".aux"): 
                    elpath = os.path.join(root, name)
                    try:
                        os.remove(elpath)
                    except:
                        print("no pude borrar "+name)
        self.progressBar.hide()

    def cualCarpeta(self):
        #self.folder.hide()
        file_dialog = getFilesDlg()
        file_dialog.sendPaths.connect(self.procesaCarpetas)
        file_dialog.exec()

    def quitaColumnaVacias(self,ultima,primera,tabName):
        for columna in range(ultima,primera,-1):
            suma = 0
            for renglon in range(0,self.tables[tabName].rowCount()):
                try:
                    suma += float(self.tables[tabName].item(renglon,columna).text())
                except:
                    print("estaba vacio")

            if suma < 0.00000001:
                self.tables[tabName].removeColumn(columna)

    def cargaCategorias(self):
        
        try:
            self.cliente_path = os.path.split(os.path.split(self.paths[0])[0])[0]
            self.folder_year = os.path.split(self.paths[0])[0]
        except:
            print("vengo de escoger cliente")

        
        self.json_path = join(self.cliente_path, "categorias_dicc_huiini.json")
        if os.path.exists(self.json_path):
            with open(self.json_path, "r", encoding="utf-8") as jsonfile:
                self.dicc_de_categorias = json.load(jsonfile)
        else:
            self.dicc_de_categorias = {}

        self.lista_categorias_default = []

        with open(join(scriptDirectory,"categorias_default.json"), "r", encoding="utf-8") as jf:
            self.lista_categorias_default = json.load(jf)
        for categoria, claves in self.dicc_de_categorias.items():
            if not categoria in self.lista_categorias_default:
                self.lista_categorias_default.append(categoria)

    def setup_log(self):
        try:
            self.pluma = open(join(self.year_folder, "log.txt"), "w")
            hoy = str(datetime.now()).split(".")[0].replace(" ","T").replace("-","_").replace(":","_")
            self.pluma.write("Log generado"+ hoy)
            self.pluma.write("\n")
            self.pluma.write("Carpetas procesadas:")
            self.pluma.write("\n")
            for path in self.paths:
                self.pluma.write(path)
                self.pluma.write("\n")
        except:
            print("no pude escribir en el log")


    def close_log(self):
        self.pluma.close()       

    def procesaCarpetas(self,paths):
        self.paths = paths.copy()
        self.progressBar.show()
        self.progressBar.setValue(1)

        self.tabWidget.clear()

        self.cargaCategorias()
        self.despliega_cliente()

        self.conceptos = []
        self.yaEstaba = {}
        self.complementosDePago = {}
        self.meses = []
        print(paths[0])
        self.year_folder = os.path.split(paths[0])[0]
        self.setup_log()
        same_year = True
        no_son_meses = False

        huiini_home_folder = os.path.split(os.path.split(self.year_folder)[0])[0]
        print(huiini_home_folder)
        with open(os.path.join(appDataDirectory,"huiini_home_folder_path.txt"), "w") as f:
            f.write(huiini_home_folder)
        meses = []
        for path in paths:
            mes = os.path.split(path)[1]
            mes = ''.join([i for i in mes if not i.isdigit()])
            mes = mes.strip()
            mes = mes.upper()
            meses.append(mes)
            if not mes in self.todos_los_meses:
                no_son_meses = True
            if self.year_folder != os.path.split(path)[0]:
                same_year = False

        self.paths = [tuple[1] for x in self.todos_los_meses for tuple in zip(meses,self.paths) if tuple[0] == x]

        print(self.paths)

        if not same_year or no_son_meses:
            if no_son_meses:
                QMessageBox.information(self, "Información", "no son meses")
            else:
                QMessageBox.information(self, "Información", "no son el mismo año")
        else:
            print(self.year_folder)
            year = os.path.split(os.path.split(paths[0])[0])[1]
            client = os.path.split(os.path.split(os.path.split(paths[0])[0])[0])[1]
            print(client+"_"+year)
            self.respaldo_anual = False
            if self.tiene_pdflatex:
                reply = QMessageBox.question(self, 'Message',"Crear pdfs?", QMessageBox.Yes |
                QMessageBox.No, QMessageBox.No)

                if reply == QMessageBox.Yes:
                    self.hacerPDFs = True
                else:
                    self.hacerPDFs = False
            else:
                self.hacerPDFs = False

            self.annual_xlsx_path = os.path.join(self.year_folder, client+"_"+year + ".xlsx")
            if os.path.isfile(self.annual_xlsx_path):#borra el anterior

                reply = QMessageBox.question(self, 'Message',"Borrar información previa?", QMessageBox.Yes |
                QMessageBox.No, QMessageBox.No)

                if reply == QMessageBox.Yes:
                    self.respaldo_anual = True
                    hoy = str(datetime.now()).split(".")[0].replace(" ","T").replace("-","_").replace(":","_")
                    self.respaldo_anual_path = self.annual_xlsx_path.split(".xlsx")[0]+"respaldo_"+hoy+".xlsx"
                    os.rename(self.annual_xlsx_path,self.respaldo_anual_path)

            cliente = os.path.split(self.cliente_path)[1]
            self.excel_path = join(self.paths[0],"EGRESOS","huiini","resumen_" + cliente + "_" + self.mes + ".xlsx")

            self.listaDeFacturasIngresos = []

            p = 0
            for path in self.paths:
                p += 1
                progreso = int(100*(p/(len(self.paths)+2)))
                self.procesaEgresos(path)
                self.pon_categorias_custom_por_factura(self.paths)
                self.pon_categorias_custom_en_gui(self.paths)
                self.quitaColumnaVacias(14,5,self.mes)
                self.agregaMes(self.mes)
                self.procesaIngresos(path)
                self.aislaNomina(path)
                self.aislaReconocibles(path)
                self.progressBar.setValue(progreso)
                if p == 1:
                    self.tabWidget.show()
                #     self.tabWidget.removeTab(0)
                #     self.tabWidget.removeTab(0)


            self.hazTabDeIngresos(self.paths)
            self.hazWsRegimenes()
            p += 1
            progreso = int(100*(p/(len(paths)+2)))
            self.progressBar.setValue(progreso)
            self.pon_categorias_custom_por_concepto(self.paths)
            self.hazAgregados(self.paths)
            p += 1
            progreso = int(100*(p/(len(paths)+2)))
            self.progressBar.setValue(progreso)

            if len(self.listaDeFacturasIngresos) > 0:
                self.agregaTab("Ingresos")
                self.quitaColumnaVacias(12,6,"Ingresos")
            #self.agregaTab("Conceptos")
            #self.agregaTab("IVA_anual")
            #self.agregaTab("Importe_anual")



            self.action_editar_Categor_as.setEnabled(True)
            self.actionClaves.setEnabled(True)
            if self.tiene_gswin64c == True:
                self.actionImprimir.setEnabled(True)
            self.excel_anual_button.setEnabled(True)
            self.raise_()
            self.activateWindow()
            self.progressBar.hide()
            if self.respaldo_anual:
                if filecmp.cmp(self.annual_xlsx_path,self.respaldo_anual_path):
                    os.remove(self.respaldo_anual_path)
                else:
                    print("se respaldó el archivo existente y el nuevo no es igual")
            curr_index = self.tabWidget.currentIndex() 
            name = self.tabWidget.tabText(curr_index)
            self.mes = name     
            self.sumale()


        if name == "Ingresos":
            print("aquí haría algo")
        else:
            self.mes = name

        self.close_log()

    




    def agregaTab(self, tabName):

        # or



        workbook = load_workbook(self.annual_xlsx_path, data_only=True)
        ws = workbook[tabName]
        c_max = ws.max_column - 9
        r_max = 0
        for r in range(9, ws.max_row+1):
            if ws.cell(r,1).value == None:
                r_max = r
                break


        self.tables[tabName] = QTableWidget(r_max,c_max)
        self.tabWidget.addTab(self.tables[tabName], tabName)
        self.tables[tabName].setColumnCount(c_max-1)

        self.tables[tabName].verticalHeader().setFixedWidth(35)
        header = self.tables[tabName].verticalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self.handleHeaderMenu)

        lc = []
        for cell in ws[8]:
            lc.append(cell.value)

        self.ponEncabezado(lc,tabName)

        # self.tables[self.mes].cellDoubleClicked.connect(self.meDoblePicaronXML)
        # self.tables[self.mes].horizontalHeader().sectionClicked.connect(self.reordena)
        for r in range(9, r_max):
            for c in range(1, c_max):
                #print(str(r),str(c))
                valor = ws.cell(r,c).value
                self.tables[tabName].setItem(r-9,c-1,self.esteItem(str(valor),str(valor)))

    def pon_categorias_custom_por_factura(self,paths):

        folder_cliente = os.path.split(os.path.split(paths[0])[0])[0]
        json_path = join(folder_cliente,"categorias_dicc_huiini.json")
        hay_categorias_custom = False
        if os.path.exists(json_path):
            hay_categorias_custom = True
            with open(json_path, "r") as jsonfile:
                self.dicc_de_categorias = json.load(jsonfile)


        if hay_categorias_custom:
            for factura in self.facturas[self.mes]:
                for concepto in factura.conceptos:
                    clave = concepto["clave_concepto"]

                    for categoria, claves in self.dicc_de_categorias.items():
                        for clave1 in claves:
                            if clave.startswith(clave1):
                                concepto["tipo"] = categoria


    def pon_categorias_custom_por_concepto(self,paths):
        folder_cliente = os.path.split(os.path.split(paths[0])[0])[0]
        json_path = join(folder_cliente,"categorias_dicc_huiini.json")
        hay_categorias_custom = False
        if os.path.exists(json_path):
            hay_categorias_custom = True
            with open(json_path, "r") as jsonfile:
                self.dicc_de_categorias = json.load(jsonfile)


        if hay_categorias_custom:
            for concepto in self.conceptos:
                clave = concepto["clave_concepto"]
                for categoria, claves in self.dicc_de_categorias.items():
                    for clave1 in claves:
                        if clave.startswith(clave1):
                            concepto["tipo"] = categoria

    def pon_categorias_custom_en_gui(self, paths):

        folder_cliente = os.path.split(os.path.split(paths[0])[0])[0]
        json_path = join(folder_cliente,"categorias_dicc_huiini.json")
        hay_categorias_custom = False
        if os.path.exists(json_path):
            hay_categorias_custom = True
            with open(json_path, "r") as jsonfile:
                self.dicc_de_categorias = json.load(jsonfile)


        if hay_categorias_custom:
            r = 0
            for factura in self.facturas[self.mes]:
                
                tooltipTipo = "\n".join(x['tipo'] for x in factura.conceptos)
                self.tables[self.mes].setItem(r,17,self.esteItem(factura.conceptos[0]['tipo'],tooltipTipo))
                r+=1

    def extrae_mes(self, folder_path):
        mes = os.path.split(folder_path)[1]
        mes = ''.join([i for i in mes if not i.isdigit()])
        mes = mes.strip()
        mes = mes.upper()
        return mes

    def procesaIngresos(self, path):
        self.esteFolder = join(path,"INGRESOS")
        if os.path.isdir(self.esteFolder):
            if not os.path.exists(join(self.esteFolder, "huiini")):
                os.makedirs(join(self.esteFolder, "huiini"))
            # self.mes = os.path.split(path)[1]
            # self.mes = ''.join([i for i in self.mes if not i.isdigit()])
            # self.mes = self.mes.strip()
            self.mes = self.extrae_mes(path)    
            cuantosDuplicados = 0
            self.listaDeDuplicados = []
            self.listaDeFacturas = []
            self.listaDeUUIDs = []
            contador = 0
            for archivo in os.listdir(self.esteFolder):
                if archivo.endswith(".xml"):
                    try:
                        laFactura = Factura(join(self.esteFolder + os.sep,archivo))
                        if laFactura.sello == "SinSello":
                            print("Omitiendo xml sin sello "+laFactura.xml_path)
                        else:
                            if laFactura.version:
                                if laFactura.UUID in self.listaDeUUIDs:

                                    cuantosDuplicados+=1
                                    self.listaDeDuplicados.append(laFactura.UUID)
                                else:
                                    self.listaDeUUIDs.append(laFactura.UUID)
                                    contador += 1
                                    self.listaDeFacturas.append(laFactura)
                    except:
                        self.warning(self, "Information", "La factura : " + join(self.esteFolder + os.sep,archivo) + " está corrupta")

            self.facturas[self.mes] = sorted(self.listaDeFacturas, key=lambda listaDeFacturas: listaDeFacturas.fechaTimbrado)
            if cuantosDuplicados > 0:
                mensaje = "En ingresos hay "+str(cuantosDuplicados)+" duplicados\n"
                chunks = []
                for esteDuplicado in self.listaDeDuplicados:
                    chunks.append(str(esteDuplicado)+"\n")
                mensaje2 = "".join(chunks)
                mensaje = mensaje + mensaje2
                self.warning(self, "Information", mensaje)

            # for t in range(0,5):
            #     time_old.sleep(0.05*len(self.facturas[self.mes]))
            #     self.pd.setValue(self.pd.value() + ( (100 - self.pd.value()) / 2))
            contador = 0

            los_facturas = self.facturas[self.mes].copy()
            self.listaDeFacturasIngresos.extend(los_facturas)
            for factura in self.facturas[self.mes]:
                #self.pd.setValue(50*((contador + 1)/len(self.facturas[self.mes])))
                factura.setFolio(contador + 1)
                if factura.tipoDeComprobante == "P":
                    if factura.IdDocumento in self.complementosDePago:
                        self.complementosDePago[factura.IdDocumento]["suma"] += factura.ImpPagado
                    else:
                        self.complementosDePago[factura.IdDocumento] = {}
                        self.complementosDePago[factura.IdDocumento]["suma"] = factura.ImpPagado

                    self.complementosDePago[factura.IdDocumento]["fechaUltimoPago"] = factura.fechaTimbrado

            if self.hacerPDFs:
                self.hazPDFs()
                #time_old.sleep(0.1*len(self.facturas[self.mes]))
                self.borraAuxiliares()

    def aislaNomina(self, path):
        esteFolder = join(path,"EGRESOS")
        # esteFolderIngresos = join(path,"INGRESOS")
        # if not os.path.exists(esteFolderIngresos):
        #     os.makedirs(esteFolderIngresos)

        for archivo in os.listdir(esteFolder):
            if archivo.endswith(".xml"):
                try:
                    laFactura = Factura(join(esteFolder + os.sep,archivo))
                    if laFactura.tipoDeComprobante == "N":
                        if not os.path.exists(join(esteFolder, "Nomina")):
                            os.makedirs(join(esteFolder, "Nomina"))
                        try:
                            os.rename(join(esteFolder + os.sep,archivo), join(esteFolder, "Nomina",archivo))

                        except:
                            print("no pude mover una nómina")
                except:
                    print("--------------------------------------------------")

    def aislaReconocibles(self, path):
        esteFolder = join(path,"EGRESOS")
        # esteFolderIngresos = join(path,"INGRESOS")
        # if not os.path.exists(esteFolderIngresos):
        #     os.makedirs(esteFolderIngresos)

        for archivo in os.listdir(esteFolder):
            if archivo.endswith(".xml"):
                try:
                    laFactura = Factura(join(esteFolder + os.sep,archivo))
                    if laFactura.tipoDeComprobante == "P":
                        if not os.path.exists(join(esteFolder, "COMPLEMENTO_PAGO")):
                            os.makedirs(join(esteFolder, "COMPLEMENTO_PAGO"))
                        try:
                            os.rename(join(esteFolder + os.sep,archivo), join(esteFolder, "COMPLEMENTO_PAGO",archivo))

                        except:
                            print("no pude mover una COMPLEMENTO_PAGO")

                    if laFactura.tipoDeComprobante == "E":
                        if not os.path.exists(join(esteFolder, "INGRESOS_NEGATIVOS")):
                            os.makedirs(join(esteFolder, "INGRESOS_NEGATIVOS"))
                        try:
                            os.rename(join(esteFolder + os.sep,archivo), join(esteFolder, "INGRESOS_NEGATIVOS",archivo))

                        except:
                            print("no pude mover una INGRESOS_NEGATIVOS")

                    if laFactura.conceptos[0]["clave_concepto"] == "78111808":
                        if not os.path.exists(join(esteFolder, "UBER")):
                            os.makedirs(join(esteFolder, "UBER"))
                        try:
                            os.rename(join(esteFolder + os.sep,archivo), join(esteFolder, "UBER",archivo))

                        except:
                            print("no pude mover una UBER")

                    
                except:
                    print("--------------------------------------------------")

    def mueveAcarpetaCoi(self, xml_path, renglon, sender):
        print("me pico ",renglon,xml_path,sender.currentText())
        xml_dir, xml_name= os.path.split(xml_path)
        print("el xml es", xml_name)
        if os.path.split(xml_dir)[1] == "EGRESOS":
            mes_folder = xml_dir
        else:
            mes_folder = os.path.split(xml_dir)[0]
        print("el folder de egresos es ", mes_folder)
        
        for (dirpath, dirnames, filenames) in os.walk(mes_folder):
            for filename in filenames:
                if filename == xml_name: 
                    print("estaba en ", join(dirpath, filename), "y lo movere a ", join(mes_folder, sender.currentText()))
                    if sender.currentText() == "--":
                        try:
                            os.rename(join(dirpath, filename), join(mes_folder, filename))

                        except:
                            print("no pude mover el xml : ",filename ," a ",mes_folder)
                    else:
                        if not os.path.exists(join(mes_folder, sender.currentText())):
                            os.makedirs(join(mes_folder, sender.currentText()))
                        try:
                            os.rename(join(dirpath, filename), join(mes_folder, sender.currentText(),filename))

                        except:
                            print("no pude mover el xml : ",filename ," a ",sender.currentText())




    def rellena_mes_gui_from_excel(self):
        workbook = load_workbook(self.annual_xlsx_path)
        if not self.mes in workbook.sheetnames:
            print("Ese mes no está:",self.mes)
        else:

            conceptos = {}
            ws_conceptos = workbook["Conceptos"]
            rows = ws_conceptos.rows
            for row in rows:
                print(row[3].value)
                if row[3].value in conceptos:
                    conceptos[row[3].value] +=  u'\n' + row[2].value
                else:
                    conceptos[row[3].value] = row[2].value

            self.tables[self.mes] = QTableWidget(10,20)
            self.setupTabMeses()
            self.tabWidget.addTab(self.tables[self.mes], self.mes)
            self.tables[self.mes].clear()
            lc = ["Pdf","Fecha","UUID","Receptor","Emisor","Concepto","Subtotal","Descuento","Traslado\nIVA","Traslado\nIEPS","Retención\nIVA","Retención\nISR","Total","Forma\nPago","Método\nPago","Tipo","Carpeta Coi"]
            self.ponEncabezado(lc,self.mes)
            self.tables[self.mes].setRowCount(13)
            self.tables[self.mes].repaint()
            self.delegate = PaddingDelegate()
            self.tables[self.mes].setItemDelegate(self.delegate)
            contador = -2
            ws_mes = workbook[self.mes]
            rows = ws_mes.rows
            for row in rows:
                contador += 1
                if contador > -1:
                    if contador > 13:
                        self.tables[self.mes].setRowCount(contador)
                    self.tables[self.mes].setItem(contador,1,self.esteItem(row[1].value,row[1].value))
                    self.tables[self.mes].setItem(contador,2,self.esteItem(row[2].value,row[2].value))
                    self.tables[self.mes].setItem(contador,3,self.esteItem("receptorRFC","receptorNombre"))
                    self.tables[self.mes].setItem(contador,4,self.esteItem(row[4].value,row[3].value))
                    message = ""
                    try:
                        message = conceptos[row[2].value]
                    except:
                        print("no tiene conceptos")
                    self.tables[self.mes].setItem(contador,5, self.esteItem(row[5].value,message))
                    self.tables[self.mes].setItem(contador,6,self.esteCenteredItem(str(row[6].value),""))
                    self.tables[self.mes].setItem(contador,7,self.esteCenteredItem(str(row[7].value),""))
                    self.tables[self.mes].setItem(contador,8,self.esteCenteredItem(str(row[8].value),""))
                    self.tables[self.mes].setItem(contador,9,self.esteCenteredItem(str(row[11].value),""))
                    self.tables[self.mes].setItem(contador,10,self.esteCenteredItem("0",""))
                    self.tables[self.mes].setItem(contador,11,self.esteCenteredItem("0",""))
                    self.tables[self.mes].setItem(contador,12,self.esteCenteredItem(str(row[12].value),""))
                    #self.tables[self.mes].setItem(contador,12,self.esteItem(row[12].value,""))
                    self.tables[self.mes].setItem(contador,13, self.esteItem(row[13].value,row[13].value))
                    self.tables[self.mes].setItem(contador,14, self.esteItem(row[14].value,row[14].value))
                    # tooltipTipo = "\n".join(x['tipo'] for x in factura.conceptos)
                    tooltipTipo = row[14].value
                    self.tables[self.mes].setItem(contador,15, self.esteItem(row[15].value,tooltipTipo))
                    # self.tables[self.mes].setCellWidget(contador,16,listacombos[contador])
                    pdf_dir = os.path.join(self.esteFolder,"huiini")
                    pdf_name = row[2].value+".pdf"
                    pdf_path = os.path.join(pdf_dir, pdf_name)
                    if os.path.exists(pdf_path):
                        self.tables[self.mes].setCellWidget(contador,0, ImgWidgetPalomita(self))

    def procesaEgresos(self, path):
        #self.folder.setText("Procesando: " + u'\n' + path)
        #self.folder.show()
        self.esteFolder = join(path,"EGRESOS")
        folder_cliente = os.path.split(os.path.split(path)[0])[0]
        coi_json_path = join(folder_cliente, "carpetas_coi.json")
        if os.path.exists(coi_json_path):
            with open(coi_json_path, "r", encoding="utf-8") as jsonfile:
                lista_carpetas_coi = json.load(jsonfile)
        else:
            lista_carpetas_coi = ["--","UBER","Nomina","COMPLEMENTO_PAGO","INGRESOS_NEGATIVOS"]
        # self.mes = os.path.split(path)[1]
        # self.mes = ''.join([i for i in self.mes if not i.isdigit()])
        # self.mes = self.mes.strip()
        self.mes = self.extrae_mes(path)
        self.meses.append(self.mes) # aqui si self.mes no se puede formar correctamantre deberia dar un error y ventanita que lo explique
        #self.conceptos[self.mes] = []
        self.tables[self.mes] = QTableWidget(10,20)
        self.setupTabMeses()
        self.tabWidget.addTab(self.tables[self.mes], self.mes)

        if not os.path.exists(join(self.esteFolder, "huiini")):
            os.makedirs(join(self.esteFolder, "huiini"))
        self.tables[self.mes].clear()
        lc = ["Pdf","Fecha","UUID","Receptor","Emisor","Concepto","Subtotal","Descuento","Traslado\nIVA","Traslado\nIEPS","Traslado\nISH","Traslado\nTUA","Retención\nIVA","Retención\nISR","Total","Forma\nPago","Método\nPago","Tipo","Carpeta Coi"]
        self.ponEncabezado(lc,self.mes)
        self.tables[self.mes].setRowCount(13)
        self.tables[self.mes].repaint()
        self.delegate = PaddingDelegate()
        self.tables[self.mes].setItemDelegate(self.delegate)
        cuantosDuplicados = 0
        self.listaDeDuplicados= []
        self.listaDeFacturas = []
        self.listaDeUUIDs = []

        listaDePathsXMLS = []
        for root, dirs, files in os.walk(self.esteFolder, topdown=False):
            for name in files:
                if name.endswith(".xml"):
                    listaDePathsXMLS.append(os.path.join(root, name))
           
        contador = 0
        for xml_path in listaDePathsXMLS:
            try:
                print(xml_path)
                laFactura = Factura(xml_path)
                if laFactura.sello == "SinSello":
                    print("Omitiendo xml sin sello "+laFactura.xml_path)
                else:
                    if laFactura.version:
                        if laFactura.UUID in self.listaDeUUIDs:

                            cuantosDuplicados+=1
                            self.listaDeDuplicados.append(laFactura.UUID)
                        else:
                            #if laFactura.tipoDeComprobante != "N":
                            self.listaDeUUIDs.append(laFactura.UUID)
                            contador += 1
                            self.listaDeFacturas.append(laFactura)
            except:
                self.warning(self, "Information", "El xml " + xml_path + " no está bien formado")
                print("El xml " + xml_path + " no está bien formado")

        if contador > 13:
            self.tables[self.mes].setRowCount(contador)

 
        
        self.facturas[self.mes] = sorted(self.listaDeFacturas, key=lambda listaDeFacturas: listaDeFacturas.fechaTimbrado)
        
        
        print(self.facturas[self.mes])



        if cuantosDuplicados > 0:
            mensaje = "En egresos de " + self.mes + " hay "+str(cuantosDuplicados)+" duplicados\n"
            chunks = []
            for esteDuplicado in self.listaDeDuplicados:
                chunks.append(str(esteDuplicado)+"\n")
            mensaje2 = "".join(chunks)
            mensaje = mensaje + mensaje2
            self.warning(self, "Information", mensaje)

        contador = 0
          
        listacombos = []
        self.sumaRFC[self.mes] = {}
        for factura in self.facturas[self.mes]:
            listacombos.append(QComboBox())
            listacombos[contador].addItems (lista_carpetas_coi)
            xml_dir, xml_name= os.path.split(factura.xml_path)
            
            if os.path.split(xml_dir)[1] == "EGRESOS":
                mes_folder = xml_dir
            else:
                mes_folder, carpeta_actual = os.path.split(xml_dir)
                listacombos[contador].setCurrentText(carpeta_actual)
    
            #listacombos[contador].setStyleSheet("QComboBox{margin:3px};")
            listacombos[contador].currentIndexChanged.connect(lambda state, xml_path=factura.xml_path, renglon=contador, combo=listacombos[contador] : self.mueveAcarpetaCoi(xml_path,renglon,combo))
            factura.setFolio(contador + 1)
            los_conceptos = factura.conceptos.copy()
            for concepto in los_conceptos:
                concepto["subTotal"] = 0.0
                concepto["mes"] = self.mes
                concepto["UUID"] = factura.UUID
                if concepto["impuestos"]:
                    concepto["impuestos"] = float(concepto['impuestos'])
                else:
                    concepto["impuestos"] = 0
                if factura.tipoDeComprobante == "E":
                    #concepto["importeConcepto"] = 0.0 - float(concepto['importeConcepto'])
                    concepto["subTotal"] = float(concepto['importeConcepto']) - float(concepto['descuento'])
                    concepto["descuento"] = 0.0 - float(concepto['descuento'])
                    concepto["impuestos"] = 0.0 - float(concepto['impuestos'])
                    #concepto["total"] = 0.0 - float(concepto['total'])

            if factura.tipoDeComprobante != "N":
                self.conceptos.extend(los_conceptos)


            if factura.tipoDeComprobante == "P":
                if factura.IdDocumento in self.complementosDePago:
                    self.complementosDePago[factura.IdDocumento]["suma"] += factura.ImpPagado
                else:
                    self.complementosDePago[factura.IdDocumento] = {}
                    self.complementosDePago[factura.IdDocumento]["suma"] = factura.ImpPagado

                self.complementosDePago[factura.IdDocumento]["fechaUltimoPago"] = factura.fechaTimbrado

            if factura.tipoDeComprobante == "N":
                self.listaDeFacturasIngresos.append(factura)

            xml_path = factura.xml_path

            files = {'files': open(xml_path , 'rb')}

            self.tables[self.mes].setItem(contador,1,self.esteItem(factura.fechaTimbrado,factura.fechaTimbrado))
            self.tables[self.mes].setItem(contador,2,self.esteItem(factura.UUID,factura.UUID))
            self.tables[self.mes].setItem(contador,3,self.esteItem(factura.ReceptorRFC,factura.ReceptorNombre))
            self.tables[self.mes].setItem(contador,4,self.esteItem(factura.EmisorRFC,factura.EmisorNombre))
            mesage = ""
            for concepto in factura.conceptos:
                try:
                    clave = concepto["clave_concepto"].strip()
                    mesage += self.concepto[clave] + u'\n'
                except:
                    self.warning(self, "Clave incorrecta", "Clave de concepto no válida en la factura" + factura.xml_path)
                    
            self.tables[self.mes].setItem(contador,5, self.esteItem(factura.conceptos[0]['descripcion'],mesage))
            self.tables[self.mes].setItem(contador,6,self.esteCenteredItem(str(factura.subTotal),""))
            self.tables[self.mes].setItem(contador,7,self.esteCenteredItem(str(factura.descuento),""))
            self.tables[self.mes].setItem(contador,8,self.esteCenteredItem(str(factura.traslados["IVA"]["importe"]),""))
            self.tables[self.mes].setItem(contador,9,self.esteCenteredItem(str(factura.traslados["IEPS"]["importe"]),""))

            self.tables[self.mes].setItem(contador,10,self.esteCenteredItem(str(factura.traslados["ISH"]["importe"]),""))
            self.tables[self.mes].setItem(contador,11,self.esteCenteredItem(str(factura.traslados["TUA"]["importe"]),""))

            self.tables[self.mes].setItem(contador,12,self.esteCenteredItem(str(factura.retenciones["IVA"]),""))
            self.tables[self.mes].setItem(contador,13,self.esteCenteredItem(str(factura.retenciones["ISR"]),""))
            self.tables[self.mes].setItem(contador,14,self.esteCenteredItem(str(factura.total),""))
            self.tables[self.mes].setItem(contador,15,self.esteItem(factura.formaDePagoStr,""))
            self.tables[self.mes].setItem(contador,16, self.esteItem(factura.metodoDePago,factura.metodoDePago))
            tooltipTipo = "\n".join(x['tipo'] for x in factura.conceptos)
            self.tables[self.mes].setItem(contador,17, self.esteItem(factura.conceptos[0]['tipo'],tooltipTipo))
            self.tables[self.mes].setCellWidget(contador,18,listacombos[contador])
           
            pdf_dir = os.path.join(self.esteFolder,"huiini")
            pdf_name = os.path.split(factura.tex_path)[1].replace("tex","pdf")
            pdf_path = os.path.join(pdf_dir, pdf_name)
            if os.path.exists(pdf_path):
                self.tables[self.mes].setCellWidget(contador,0, ImgWidgetPalomita(self))
            
            if factura.EmisorRFC in self.sumaRFC[self.mes]:
                self.sumaRFC[self.mes][factura.EmisorRFC]['subTotal'] += float(factura.subTotal)
                self.sumaRFC[self.mes][factura.EmisorRFC]['descuento'] += float(factura.descuento)
                self.sumaRFC[self.mes][factura.EmisorRFC]['trasladoIVA'] += float(factura.traslados['IVA']['importe'])
                self.sumaRFC[self.mes][factura.EmisorRFC]['importe'] += float(factura.subTotal)-float(factura.descuento)
                self.sumaRFC[self.mes][factura.EmisorRFC]['total'] += float(factura.total)
                self.sumaRFC[self.mes][factura.EmisorRFC]['importeStr'] += "+"+str(float(factura.subTotal)-float(factura.descuento))
                self.sumaRFC[self.mes][factura.EmisorRFC]['trasladoIVAStr'] += "+"+str(factura.traslados['IVA']['importe'])
                print("sumale " + str(factura.subTotal) )
            else:
                self.sumaRFC[self.mes][factura.EmisorRFC] = {'subTotal': float(factura.subTotal),
                                                              'descuento': float(factura.descuento),
                                                              'trasladoIVA': float(factura.traslados['IVA']['importe']),
                                                              'importe': float(factura.subTotal)-float(factura.descuento),
                                                              'total': float(factura.total),
                                                              'importeStr': "="+str(float(factura.subTotal)-float(factura.descuento)),
                                                              'trasladoIVAStr': "="+str(factura.traslados['IVA']['importe']),
                                                              'nombre': factura.EmisorNombre
                                                            }
                print("crealo con " + str(factura.subTotal))

            contador += 1
        

        if self.hacerPDFs:
            self.hazPDFs()
            #time_old.sleep(0.2*len(self.facturas[self.mes]))
            self.borraAuxiliares()


        contador = -1

        # time_old.sleep(0.5*len(self.facturas[self.mes]))

        self.numeroDeFacturasValidas[self.mes] = len(self.facturas[self.mes])


        self.sumale()

        self.hazResumenDiot(self.esteFolder)
        #if len(paths)>2:
        #obtener los warnings de las facturas
        mensajeAlerta =""
        for factura in self.facturas[self.mes]:
            if not factura.mensaje == "":
                mensajeAlerta += factura.UUID + factura.mensaje + r'\n'
        if not mensajeAlerta == "":
            self.warning(self, "Information", mensajeAlerta)

        


        #self.folder.setText("Carpeta Procesada: " + u'\n' + self.esteFolder)
        #self.folder.show()

        

class WheelEventFilter(QtCore.QObject):
    def eventFilter(self, obj, ev):
        if obj.inherits("QComboBox") and ev.type() == QtCore.QEvent.Wheel:
            return True
        return False
class PaddingDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, padding=5, parent=None):
        super(PaddingDelegate, self).__init__(parent)
        self._padding = ' ' * max(1, padding)


    def is_num(self,s):
        try:
            float(s.replace(',',''))
        except ValueError:
            return False
        else:
            return True

    def displayText(self, text, locale):
        if self.is_num(text):
            return text + self._padding
        else:
            return self._padding + text

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        margins = editor.textMargins()
        padding = editor.fontMetrics().width(self._padding) + 1
        margins.setLeft(margins.left() + padding)
        editor.setTextMargins(margins)
        return editor




#app = QtWidgets.QApplication.instance()
app = QtWidgets.QApplication(sys.argv)
filter = WheelEventFilter()
app.installEventFilter(filter)
app.setStyleSheet("QMessageBox { messagebox-text-interaction-flags: 5; }")
form = Ui_MainWindow()
form.show()


app.exec_()




# 29fd64f4-f6d5-4b9a-b768-819a2ac38589 no pudo sumar un traslado local