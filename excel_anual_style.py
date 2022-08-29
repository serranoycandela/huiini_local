from openpyxl import load_workbook,  Workbook
import openpyxl
from openpyxl.styles.alignment import Alignment
import datetime
workbook = Workbook()
ws_mes = workbook.create_sheet("ENERO")
img = openpyxl.drawing.image.Image('C:\\Users\\serra\\GitHub\\huiini_local\\logo_s.png')
  
# The Coordinates where the image would be pasted
# (an image could span several rows and columns
# depending on it's size)
img.anchor = 'A2'
  
# Adding the image to the worksheet
# (with attributes like position)
ws_mes.add_image(img)

ws_mes.cell(2, 4, "Nombre: ")

ws_mes.cell(3, 4, "RFC: ")
ws_mes.cell(4, 4, "Periodo: ")
ws_mes.cell(5, 4, "Contador: ")
ws_mes.cell(6, 4, "Fecha de Actualización: ") 
ws_mes.cell(6, 5, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

for row in ws_mes["D2":"D6"]:
    for cell in row:
        print(cell.value)
        cell.alignment = Alignment(horizontal="right")

ws_mes.column_dimensions['A'].width = 30

img = openpyxl.drawing.image.Image('C:\\Users\\serra\\GitHub\\huiini_local\\logo.png')
  
# The Coordinates where the image would be pasted
# (an image could span several rows and columns
# depending on it's size)
img.anchor = 'O2'
  
# Adding the image to the worksheet
# (with attributes like position)
ws_mes.add_image(img)
  
# Saving the workbook created under the name of out.xlsx
workbook.save('C:/Users/serra/GitHub/huiini_local/out.xlsx')